#!/usr/bin/env python3
"""Processing — logica di calcolo + generazione output Excel."""

import io
import os
import re
import sys
import tempfile
import streamlit as st
import pandas as pd
from datetime import datetime
from typing import Tuple, Dict, Set, Optional

from tour_operators import (
    detect_tour_operators,
    find_tour_operator_folder,
    get_tour_operator_module_name,
    get_tour_operator_processors,
    load_holiday_list,
    write_output_excel_veratour,
    write_output_excel_alpitour,
    ALPITOUR_AVAILABLE,
)

# Importa extract_atd_candidates da Veratour (funzione condivisa)
# Viene iniettata nei moduli che la usano ma non la definiscono
try:
    from consuntivoveratour import extract_atd_candidates as _extract_atd_candidates
except ImportError:
    _extract_atd_candidates = None

# Moduli che supportano il nuovo formato 2026 nativamente
_NEW_FORMAT_MODULES = {'veratour', 'alpitour'}


def _make_compat_excel(input_path: str) -> Optional[str]:
    """
    Se l'Excel usa il nuovo formato 2026 (INIZIO TURNO + FINE TURNO),
    crea un file temp con:
    - colonna TURNO sintetizzata (HH:MM-HH:MM) da INIZIO+FINE dove disponibili
    - TURNO da STD-2h30→STD per righe senza INIZIO TURNO
    - INIZIO TURNO e FINE TURNO RIMOSSI → forza percorso old-format nei moduli
    Restituisce path al file temp, o None se non necessario.
    """
    try:
        xls = pd.ExcelFile(input_path)
        needs_conversion = False
        sheets_data = {}

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(input_path, sheet_name=sheet_name)
            cols_up = [str(c).strip().upper() for c in df.columns]

            has_inizio = any('INIZIO TURNO' in c for c in cols_up)
            has_fine   = any('FINE TURNO'   in c for c in cols_up)
            has_turno  = any(c in ('TURNO', 'TURNI') for c in cols_up)

            if has_inizio and has_fine and not has_turno:
                needs_conversion = True
                inizio_col = next(c for c in df.columns if 'INIZIO TURNO' in str(c).strip().upper())
                fine_col   = next(c for c in df.columns if 'FINE TURNO'   in str(c).strip().upper())
                std_col    = next((c for c in df.columns if str(c).strip().upper() == 'STD'), None)

                def _fmt(val):
                    if pd.isna(val): return ''
                    if isinstance(val, pd.Timedelta):
                        s = int(val.total_seconds())
                        return f"{s//3600:02d}:{(s%3600)//60:02d}"
                    if isinstance(val, (pd.Timestamp, datetime)):
                        return f"{val.hour:02d}:{val.minute:02d}"
                    s = str(val).strip()
                    m = re.match(r'(\d{1,2})[:\.](\d{2})', s)
                    return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}" if m else s

                turno_vals = []
                for _, row in df.iterrows():
                    ini = _fmt(row[inizio_col])
                    fin = _fmt(row[fine_col])
                    if ini and fin:
                        turno_vals.append(f"{ini}-{fin}")
                    elif std_col is not None:
                        std_v = row.get(std_col)
                        std_s = _fmt(std_v) if std_v is not None else ''
                        if std_s:
                            try:
                                sh, sm = map(int, std_s.split(':'))
                                total = (sh * 60 + sm - 150) % 1440
                                turno_vals.append(f"{total//60:02d}:{total%60:02d}-{std_s}")
                            except Exception:
                                turno_vals.append(None)
                        else:
                            turno_vals.append(None)
                    else:
                        turno_vals.append(None)

                df['TURNO'] = turno_vals
                # CHIAVE: rimuovi INIZIO TURNO e FINE TURNO → i moduli usano il vecchio percorso TURNO
                df = df.drop(columns=[inizio_col, fine_col], errors='ignore')

            sheets_data[sheet_name] = df

        if not needs_conversion:
            return None

        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return tmp_path

    except Exception:
        return None  # se fallisce, usa il file originale


def run_calculation(
    tmp_path: str,
    uploaded_file_name: str,
    apt_filter,
    night_mode: str,
    round_extra_mode: str,
    round_extra_step: int,
    round_night_mode: str,
    round_night_step: int,
    holiday_file,
) -> dict:
    """
    Esegue l'elaborazione completa.
    Restituisce dict con chiavi:
      output_buffer, output_filename, detail_df, totals_df, discr_df,
      processed_count, errors
    Esattamente come nell'originale funzionante da git main.
    """
    # Carica festivi se presente
    holiday_dates = None
    if holiday_file is not None:
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".txt") as tmp_holiday:
            tmp_holiday.write(holiday_file.getvalue().decode('utf-8'))
            tmp_holiday_path = tmp_holiday.name
        try:
            holiday_dates = load_holiday_list(tmp_holiday_path)
        finally:
            os.unlink(tmp_holiday_path)

    # Rileva tour operator
    tour_operators, aliservice_managed = detect_tour_operators(tmp_path)

    # Verifica se Aliservice è presente
    aliservice_found = False
    aliservice_folder = find_tour_operator_folder("Aliservice")
    if aliservice_managed and aliservice_folder:
        from tour_operators import ALISERVICE_AVAILABLE
        if ALISERVICE_AVAILABLE:
            aliservice_found = True

    # Filtra TO gestiti da Aliservice dalla lista principale
    tour_operators_to_check = tour_operators - aliservice_managed

    # Mappa i tour operator trovati ai loro nomi normalizzati
    # (solo prima occorrenza per modulo — usato per foglio TourOperatourRilevati)
    found_tour_operators: Dict[str, dict] = {}

    all_detail_dfs = []
    all_totals_dfs = []
    all_discr_dfs = []

    # Dizionario processori
    tour_operator_processors = get_tour_operator_processors(
        apt_filter, night_mode, round_extra_mode, round_extra_step,
        round_night_mode, round_night_step, holiday_dates,
    )

    # Prepara lista tour operator da elaborare:
    # FIX — un entry per ogni TO rilevato (no deduplication per modulo)
    # e to_keyword = nome reale del TO nel file Excel
    tour_operators_to_process = []
    warnings_list = []

    for to_name in sorted(tour_operators_to_check):
        folder_path = find_tour_operator_folder(to_name)
        if folder_path:
            module_name = get_tour_operator_module_name(to_name)
            if module_name:
                # Tieni solo la prima occorrenza per il foglio di riepilogo
                if module_name not in found_tour_operators:
                    found_tour_operators[module_name] = {
                        'original_name': to_name,
                        'folder': folder_path,
                    }
                # Aggiungi SEMPRE un entry separato per ogni TO rilevato
                if module_name in tour_operator_processors:
                    processor = tour_operator_processors[module_name]
                    if processor['available'] and processor['config_class'] and processor['process_func']:
                        tour_operators_to_process.append({
                            'name': to_name,
                            'module_name': module_name,
                            'processor': processor,
                            'to_keyword': to_name.lower(),  # FIX: usa nome reale dal file
                            'is_aliservice': False,
                        })
                    else:
                        warnings_list.append(
                            f"{to_name} rilevato ma modulo non disponibile."
                        )

    # Aggiungi Aliservice se presente
    if aliservice_found and 'aliservice' in tour_operator_processors:
        processor = tour_operator_processors['aliservice']
        if processor['available'] and processor['config_class'] and processor['process_func']:
            tour_operators_to_process.append({
                'name': 'Aliservice',
                'module_name': 'aliservice',
                'processor': processor,
                'to_keyword': 'aliservice',
                'is_aliservice': True,
            })
        else:
            warnings_list.append("Aliservice rilevato ma modulo non disponibile.")

    # Mostra warning
    for w in warnings_list:
        st.warning(w)

    # Mostra i TO che verranno elaborati
    if tour_operators_to_process:
        num_cols = min(len(tour_operators_to_process), 4)
        cols = st.columns(num_cols)
        for idx, to_info in enumerate(tour_operators_to_process):
            with cols[idx % num_cols]:
                st.info(f"Elaborazione {to_info['name']}...")

    # Elabora tutti i tour operator
    processed_count = 0
    errors = []

    for to_info in tour_operators_to_process:
        processor = to_info['processor']
        try:
            # FIX 1: inietta extract_atd_candidates nel namespace del modulo se mancante
            if _extract_atd_candidates is not None:
                module_name_str = processor['process_func'].__module__
                mod = sys.modules.get(module_name_str)
                if mod and not hasattr(mod, 'extract_atd_candidates'):
                    setattr(mod, 'extract_atd_candidates', _extract_atd_candidates)

            config_kwargs = processor['config_kwargs']()
            config_kwargs['to_keyword'] = to_info['to_keyword']
            cfg = processor['config_class'](**config_kwargs)

            # FIX 2: per moduli old-format, pre-converte Excel nuovo formato → vecchio formato
            compat_path = None
            if to_info['module_name'] not in _NEW_FORMAT_MODULES:
                compat_path = _make_compat_excel(tmp_path)

            process_path = compat_path if compat_path else tmp_path
            detail, totals, discr = processor['process_func']([process_path], cfg)

            if compat_path and os.path.exists(compat_path):
                os.unlink(compat_path)

            all_detail_dfs.append(detail)
            all_totals_dfs.append(totals)
            all_discr_dfs.append(discr)
            processed_count += 1
        except Exception as e:
            errors.append(f"{to_info['name']}: {str(e)}")

    # Mostra errori
    for error in errors:
        st.error(f"Errore durante l'elaborazione: {error}")

    if processed_count == 0:
        st.error("Nessun tour operatour con calcolo disponibile trovato nel file.")
        return None

    # Combina risultati
    if all_detail_dfs:
        detail_df = pd.concat(all_detail_dfs, ignore_index=True)
        totals_df = pd.concat(all_totals_dfs, ignore_index=True)
        discr_list = [d for d in all_discr_dfs if d is not None and not d.empty]
        discr_df = pd.concat(discr_list, ignore_index=True) if discr_list else pd.DataFrame()
    else:
        detail_df = pd.DataFrame()
        totals_df = pd.DataFrame()
        discr_df = pd.DataFrame()

    # Genera output Excel
    output_buffer = io.BytesIO()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:
        output_path = tmp_output.name

    # Usa la funzione di scrittura appropriata
    if processed_count > 1 and ALPITOUR_AVAILABLE and write_output_excel_alpitour:
        write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
    elif processed_count > 0:
        first_processed = list(found_tour_operators.keys())[0] if found_tour_operators else None
        if aliservice_found and 'aliservice' in tour_operator_processors:
            processor = tour_operator_processors['aliservice']
            if processor.get('write_func'):
                processor['write_func'](output_path, detail_df, totals_df, discr_df)
            elif ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
            else:
                write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
        elif first_processed and first_processed in tour_operator_processors:
            processor = tour_operator_processors[first_processed]
            if processor.get('write_func'):
                processor['write_func'](output_path, detail_df, totals_df, discr_df)
            elif ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
            else:
                write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
        else:
            if ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
            else:
                write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
    else:
        write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)

    # Aggiungi foglio TourOperatourRilevati
    _add_tour_operator_sheet(
        output_path, detail_df, tour_operators, aliservice_managed,
        aliservice_found, found_tour_operators, tour_operator_processors,
    )

    # Leggi il file generato
    with open(output_path, 'rb') as f:
        output_buffer.write(f.read())
    output_buffer.seek(0)

    # Cleanup
    if os.path.exists(output_path):
        os.unlink(output_path)

    return {
        'output_buffer': output_buffer.getvalue(),
        'output_filename': f"OUT_{uploaded_file_name}",
        'detail_df': detail_df,
        'totals_df': totals_df,
        'discr_df': discr_df,
        'processed_count': processed_count,
        'errors': errors,
    }


def _add_tour_operator_sheet(
    output_path: str,
    detail_df: pd.DataFrame,
    tour_operators: Set[str],
    aliservice_managed: Set[str],
    aliservice_found: bool,
    found_tour_operators: dict,
    tour_operator_processors: dict,
):
    """Aggiunge il foglio TourOperatourRilevati all'Excel di output."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(output_path)

    if "TourOperatourRilevati" in wb.sheetnames:
        wb.remove(wb["TourOperatourRilevati"])
    if "TourOperatourNonElaborati" in wb.sheetnames:
        wb.remove(wb["TourOperatourNonElaborati"])

    ws = wb.create_sheet("TourOperatourRilevati", 0)

    elaborated_tour_operators: Set[str] = set()
    if not detail_df.empty and 'TOUR OPERATOR' in detail_df.columns:
        elaborated_tour_operators = set(detail_df['TOUR OPERATOR'].dropna().astype(str).unique())

    tour_operators_for_list = tour_operators - aliservice_managed
    if aliservice_found:
        tour_operators_for_list.add("ALISERVICE")

    tour_operator_list = []
    for to_name in sorted(tour_operators_for_list):
        to_clean = re.sub(r'[^a-zA-Z]', '', to_name).lower()
        is_supported = False
        status = "Non codificato"

        if to_name.upper() == "ALISERVICE" and aliservice_found:
            if not detail_df.empty:
                if 'AGENZIA' in detail_df.columns:
                    if detail_df['AGENZIA'].astype(str).str.contains('aliservice', case=False, na=False).any():
                        is_supported = True
                        status = "Codificato - Elaborato"
                elif 'TOUR OPERATOR' in detail_df.columns:
                    for elaborated_to in elaborated_tour_operators:
                        if 'aliservice' in str(elaborated_to).lower():
                            is_supported = True
                            status = "Codificato - Elaborato"
                            break

            if not is_supported:
                is_supported = True
                status = "Codificato - Rilevato ma senza dati elaborati"
        else:
            if not detail_df.empty and 'TOUR OPERATOR' in detail_df.columns:
                for elaborated_to in elaborated_tour_operators:
                    elaborated_clean = re.sub(r'[^a-zA-Z]', '', str(elaborated_to)).lower()
                    if to_clean == elaborated_clean or to_clean in elaborated_clean or elaborated_clean in to_clean:
                        is_supported = True
                        status = "Codificato - Elaborato"
                        break

            if not is_supported:
                module_name = get_tour_operator_module_name(to_name)
                if module_name and module_name in found_tour_operators:
                    is_supported = True
                    status = "Codificato - Rilevato ma senza dati elaborati"
                elif module_name and module_name in tour_operator_processors:
                    processor = tour_operator_processors[module_name]
                    if processor['available'] and processor['config_class']:
                        is_supported = True
                        status = "Codificato - Rilevato ma senza dati elaborati"

        if not is_supported:
            folder_path = find_tour_operator_folder(to_name)
            if folder_path:
                status = "Modulo presente ma non rilevato nel file"
            else:
                status = "Non codificato"

        note = (
            "Calcolo tariffe disponibile e applicato" if status == "Codificato - Elaborato" else
            "Calcolo tariffe disponibile ma nessun dato da elaborare" if "Codificato" in status else
            "Calcolo tariffe non disponibile - da codificare"
        )

        tour_operator_list.append({
            "Tour Operatour": to_name,
            "Status": status,
            "Note": note,
        })

    # Scrivi header
    ws.cell(row=1, column=1, value="Tour Operatour")
    ws.cell(row=1, column=2, value="Status")
    ws.cell(row=1, column=3, value="Note")

    for idx, to_info in enumerate(tour_operator_list, 2):
        ws.cell(row=idx, column=1, value=to_info["Tour Operatour"])
        ws.cell(row=idx, column=2, value=to_info["Status"])
        ws.cell(row=idx, column=3, value=to_info["Note"])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60

    wb.save(output_path)
