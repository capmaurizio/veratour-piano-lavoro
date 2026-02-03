#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script per convertire file Excel in formato ottimizzato per Google Sheets
Preserva formule, colori e riferimenti
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
from pathlib import Path

def converti_formula_excel_a_google(formula_excel):
    """Converte formule Excel in formule Google Sheets dove possibile"""
    if not formula_excel or not formula_excel.startswith('='):
        return formula_excel
    
    # Alcune conversioni comuni
    # Excel usa ; come separatore, Google Sheets usa ,
    formula_google = formula_excel.replace(';', ',')
    
    # Alcune funzioni hanno nomi diversi
    replacements = {
        'CONCATENATE': 'CONCAT',
        'IFERROR': 'IFERROR',  # Stesso nome
        'VLOOKUP': 'VLOOKUP',  # Stesso nome
        'SUMIF': 'SUMIF',  # Stesso nome
        'COUNTIF': 'COUNTIF',  # Stesso nome
    }
    
    for excel_func, google_func in replacements.items():
        formula_google = formula_google.replace(excel_func, google_func)
    
    return formula_google

def preserva_stili_cella(cella_originale, cella_nuova):
    """Copia tutti gli stili dalla cella originale alla nuova"""
    try:
        # Colore di sfondo
        if cella_originale.fill and cella_originale.fill.start_color:
            if cella_originale.fill.start_color.rgb:
                cella_nuova.fill = PatternFill(
                    start_color=cella_originale.fill.start_color.rgb,
                    end_color=cella_originale.fill.end_color.rgb if cella_originale.fill.end_color and cella_originale.fill.end_color.rgb else cella_originale.fill.start_color.rgb,
                    fill_type=cella_originale.fill.fill_type or 'solid'
                )
        
        # Font
        if cella_originale.font:
            font_color = None
            if cella_originale.font.color and hasattr(cella_originale.font.color, 'rgb'):
                font_color = cella_originale.font.color.rgb
            
            cella_nuova.font = Font(
                name=cella_originale.font.name,
                size=cella_originale.font.size,
                bold=cella_originale.font.bold,
                italic=cella_originale.font.italic,
                color=font_color
            )
        
        # Allineamento
        if cella_originale.alignment:
            cella_nuova.alignment = Alignment(
                horizontal=cella_originale.alignment.horizontal,
                vertical=cella_originale.alignment.vertical,
                wrap_text=cella_originale.alignment.wrap_text,
                shrink_to_fit=cella_originale.alignment.shrink_to_fit,
                indent=cella_originale.alignment.indent
            )
        
        # Bordi - copia elemento per elemento
        if cella_originale.border:
            border_originale = cella_originale.border
            cella_nuova.border = Border(
                left=border_originale.left,
                right=border_originale.right,
                top=border_originale.top,
                bottom=border_originale.bottom,
                diagonal=border_originale.diagonal,
                diagonal_direction=border_originale.diagonal_direction,
                outline=border_originale.outline,
                vertical=border_originale.vertical,
                horizontal=border_originale.horizontal
            )
        
        # Formato numero
        if cella_originale.number_format:
            cella_nuova.number_format = cella_originale.number_format
    except Exception as e:
        # Se c'è un errore nella copia degli stili, continua comunque
        pass

def converti_file_per_google_sheets(file_input, file_output):
    """Converte un file Excel in formato ottimizzato per Google Sheets"""
    
    print(f"📖 Leggo il file: {file_input}")
    wb_originale = openpyxl.load_workbook(file_input, data_only=False)  # data_only=False per preservare formule
    
    print(f"📝 Creo nuovo file: {file_output}")
    wb_nuovo = openpyxl.Workbook()
    
    # Rimuovi il foglio di default
    if 'Sheet' in wb_nuovo.sheetnames:
        wb_nuovo.remove(wb_nuovo['Sheet'])
    
    # Copia ogni foglio
    for nome_foglio in wb_originale.sheetnames:
        print(f"  📄 Elaboro foglio: {nome_foglio}")
        ws_originale = wb_originale[nome_foglio]
        ws_nuovo = wb_nuovo.create_sheet(title=nome_foglio)
        
        # Copia tutte le celle
        for row in ws_originale.iter_rows():
            for cella_originale in row:
                if cella_originale.value is not None or cella_originale.has_style:
                    # Crea nuova cella
                    cella_nuova = ws_nuovo.cell(
                        row=cella_originale.row,
                        column=cella_originale.column
                    )
                    
                    # Copia valore o formula
                    if cella_originale.data_type == 'f':  # Formula
                        # Preserva la formula originale (Google Sheets può leggere formule Excel)
                        cella_nuova.value = cella_originale.value
                        # Opzionalmente converte in formato Google Sheets
                        # cella_nuova.value = converti_formula_excel_a_google(str(cella_originale.value))
                    else:
                        # Copia valore normale
                        cella_nuova.value = cella_originale.value
                    
                    # Preserva tutti gli stili
                    preserva_stili_cella(cella_originale, cella_nuova)
        
        # Copia larghezza colonne
        for col_letter in ws_originale.column_dimensions:
            if col_letter in ws_originale.column_dimensions:
                ws_nuovo.column_dimensions[col_letter].width = ws_originale.column_dimensions[col_letter].width
        
        # Copia altezza righe
        for row_num in ws_originale.row_dimensions:
            if row_num in ws_originale.row_dimensions:
                ws_nuovo.row_dimensions[row_num].height = ws_originale.row_dimensions[row_num].height
        
        # Copia merged cells
        for merged_range in ws_originale.merged_cells.ranges:
            ws_nuovo.merge_cells(str(merged_range))
        
        print(f"    ✅ Completato: {ws_originale.max_row} righe, {ws_originale.max_column} colonne")
    
    # Salva il nuovo file
    print(f"💾 Salvo il file: {file_output}")
    wb_nuovo.save(file_output)
    print(f"✅ Conversione completata!")
    
    return file_output

def main():
    if len(sys.argv) < 2:
        print("Uso: python converti_per_google_sheets.py <file_input.xlsx> [file_output.xlsx]")
        sys.exit(1)
    
    file_input = sys.argv[1]
    
    if not os.path.exists(file_input):
        print(f"❌ Errore: File non trovato: {file_input}")
        sys.exit(1)
    
    if len(sys.argv) >= 3:
        file_output = sys.argv[2]
    else:
        # Crea nome file output automaticamente
        path_input = Path(file_input)
        file_output = str(path_input.parent / f"{path_input.stem}_GOOGLE_SHEETS{path_input.suffix}")
    
    try:
        converti_file_per_google_sheets(file_input, file_output)
        print(f"\n📌 File convertito salvato in: {file_output}")
        print(f"📌 Ora puoi caricare questo file su Google Drive/Google Sheets")
    except Exception as e:
        print(f"❌ Errore durante la conversione: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
