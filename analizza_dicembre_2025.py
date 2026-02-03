#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('Documentazione/RIEPILOGO ASSISTENZE  BGY SORRENTI GAIA .xlsx', data_only=False)
ws = wb['DICEMBRE 2025']
print('Foglio: DICEMBRE 2025')
print(f'Righe: {ws.max_row}, Colonne: {ws.max_column}')
print('\n=== Prime 5 righe ===')
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=False), 1):
    values = []
    for cell in row[:10]:
        if cell.value:
            values.append(str(cell.value)[:30])
        else:
            values.append('')
    print(f'Riga {i}: {" | ".join(values)}')
print('\n=== Formule trovate ===')
for row in ws.iter_rows(max_row=20):
    for cell in row:
        if cell.data_type == 'f':
            print(f'{cell.coordinate}: {cell.value}')
