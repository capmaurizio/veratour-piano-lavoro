#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Interfaccia Web Streamlit Multi-Tour Operatour
Basata sulla versione originale funzionante di Veratour
"""

import streamlit as st
import pandas as pd
import io
import os
import sys
import tempfile
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set
import re

# Aggiungi Veratour, Alpitour, Aliservice, Baobab, Domina, MichelTours, SAND, Caboverdetime e Rusconi al path per import
veratour_path = os.path.join(os.path.dirname(__file__), 'Veratour')
alpitour_path = os.path.join(os.path.dirname(__file__), 'Alpitour')
aliservice_path = os.path.join(os.path.dirname(__file__), 'Aliservice')
baobab_path = os.path.join(os.path.dirname(__file__), 'Baobab')
domina_path = os.path.join(os.path.dirname(__file__), 'Domina')
micheltours_path = os.path.join(os.path.dirname(__file__), 'MICHELTOURS')
sand_path = os.path.join(os.path.dirname(__file__), ' Sand')
caboverdetime_path = os.path.join(os.path.dirname(__file__), 'Caboverdetime')
rusconi_path = os.path.join(os.path.dirname(__file__), 'Rusconi')
if veratour_path not in sys.path:
    sys.path.insert(0, veratour_path)
if alpitour_path not in sys.path:
    sys.path.insert(0, alpitour_path)
if aliservice_path not in sys.path:
    sys.path.insert(0, aliservice_path)
if baobab_path not in sys.path:
    sys.path.insert(0, baobab_path)
if domina_path not in sys.path:
    sys.path.insert(0, domina_path)
if micheltours_path not in sys.path:
    sys.path.insert(0, micheltours_path)
if sand_path not in sys.path:
    sys.path.insert(0, sand_path)
if caboverdetime_path not in sys.path:
    sys.path.insert(0, caboverdetime_path)
if rusconi_path not in sys.path:
    sys.path.insert(0, rusconi_path)

from consuntivoveratour import (
    CalcConfig as VeratourCalcConfig,
    RoundingPolicy,
    process_files as process_files_veratour,
    write_output_excel as write_output_excel_veratour,
    load_holiday_list
)

try:
    from consuntivoalpitour import (
        CalcConfig as AlpitourCalcConfig,
        process_files as process_files_alpitour,
        write_output_excel as write_output_excel_alpitour,
        validate_file_complete as validate_file_alpitour
    )
    ALPITOUR_AVAILABLE = True
except ImportError as e:
    ALPITOUR_AVAILABLE = False
    # Non mostrare warning all'avvio, solo se necessario
    AlpitourCalcConfig = None
    process_files_alpitour = None
    write_output_excel_alpitour = None
    validate_file_alpitour = None

try:
    from consuntivoaliservice import (
        CalcConfig as AliserviceCalcConfig,
        process_files as process_files_aliservice,
        write_output_excel as write_output_excel_aliservice,
    )
    ALISERVICE_AVAILABLE = True
except ImportError as e:
    ALISERVICE_AVAILABLE = False
    AliserviceCalcConfig = None
    process_files_aliservice = None
    write_output_excel_aliservice = None

try:
    from consuntivobaobab import (
        CalcConfig as BaobabCalcConfig,
        process_files as process_files_baobab,
        write_output_excel as write_output_excel_baobab,
    )
    BAOBAB_AVAILABLE = True
except ImportError as e:
    BAOBAB_AVAILABLE = False
    BaobabCalcConfig = None
    process_files_baobab = None
    write_output_excel_baobab = None

try:
    from consuntivodomina import (
        CalcConfig as DominaCalcConfig,
        process_files as process_files_domina,
        write_output_excel as write_output_excel_domina,
    )
    DOMINA_AVAILABLE = True
except ImportError as e:
    DOMINA_AVAILABLE = False
    DominaCalcConfig = None
    process_files_domina = None
    write_output_excel_domina = None

try:
    from consuntivocaboverdetime import (
        CalcConfig as CaboverdetimeCalcConfig,
        process_files as process_files_caboverdetime,
        write_output_excel as write_output_excel_caboverdetime,
    )
    CABOVERDETIME_AVAILABLE = True
except ImportError as e:
    CABOVERDETIME_AVAILABLE = False
    CaboverdetimeCalcConfig = None
    process_files_caboverdetime = None
    write_output_excel_caboverdetime = None

try:
    from consuntivorusconi import (
        CalcConfig as RusconiCalcConfig,
        process_files as process_files_rusconi,
        write_output_excel as write_output_excel_rusconi,
    )
    RUSCONI_AVAILABLE = True
except ImportError as e:
    RUSCONI_AVAILABLE = False
    RusconiCalcConfig = None
    process_files_rusconi = None
    write_output_excel_rusconi = None

try:
    from consuntivomicheltours import (
        CalcConfig as MichelToursCalcConfig,
        process_files as process_files_micheltours,
        write_output_excel as write_output_excel_micheltours,
    )
    MICHELTOURS_AVAILABLE = True
except ImportError as e:
    MICHELTOURS_AVAILABLE = False
    MichelToursCalcConfig = None
    process_files_micheltours = None
    write_output_excel_micheltours = None

try:
    from consuntivosand import (
        CalcConfig as SandCalcConfig,
        process_files as process_files_sand,
        write_output_excel as write_output_excel_sand,
    )
    SAND_AVAILABLE = True
except ImportError as e:
    SAND_AVAILABLE = False
    SandCalcConfig = None
    process_files_sand = None
    write_output_excel_sand = None

# Configurazione pagina
st.set_page_config(
    page_title="Calcolo Piano Lavoro - Multi-Tour Operatour",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizzato
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        margin-top: 1rem;
    }
    .logout-button-container {
        margin-bottom: 1rem;
    }
    .login-container {
        max-width: 450px;
        margin: 80px auto;
        padding: 40px;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        background-color: #ffffff;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .login-title {
        font-size: 1.8rem;
        font-weight: 600;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 10px;
        letter-spacing: 1px;
    }
    .login-subtitle {
        font-size: 0.95rem;
        color: #666;
        text-align: center;
        margin-bottom: 30px;
    }
    .company-name {
        font-size: 1.2rem;
        font-weight: 500;
        color: #333;
        text-align: center;
        margin-top: 20px;
        padding-top: 20px;
        border-top: 1px solid #e0e0e0;
    }
    
    /* Spaziatura per le card */
    [data-testid="column"] {
        padding: 0.5rem;
    }
    
    /* Modern button styling - Card-like buttons */
    .stButton > button {
        border-radius: 16px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        font-weight: 500;
        border: none !important;
        padding: 1.5rem !important;
        min-height: 180px !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        text-align: center !important;
        white-space: pre-line !important;
        line-height: 1.6 !important;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1), 0 2px 4px rgba(0, 0, 0, 0.06) !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-8px) scale(1.02) !important;
        box-shadow: 0 20px 25px rgba(0, 0, 0, 0.15), 0 10px 10px rgba(0, 0, 0, 0.1) !important;
    }
    
    /* Card colors based on button key */
    button[key="btn_carica"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
    }
    
    button[key="btn_calcolo"] {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%) !important;
        color: white !important;
    }
    
    button[key="btn_risultati"] {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%) !important;
        color: white !important;
    }
    
    button[key="btn_scarica"] {
        background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%) !important;
        color: white !important;
    }
    
    /* Active state (primary buttons) */
    button[key="btn_carica"][data-baseweb="button"][kind="primary"],
    button[key="btn_calcolo"][data-baseweb="button"][kind="primary"],
    button[key="btn_risultati"][data-baseweb="button"][kind="primary"],
    button[key="btn_scarica"][data-baseweb="button"][kind="primary"] {
        transform: translateY(-4px) !important;
        box-shadow: 0 12px 20px rgba(0, 0, 0, 0.2), 0 6px 8px rgba(0, 0, 0, 0.15) !important;
    }
    
    /* Secondary buttons (non-active) - slightly darker */
    button[key="btn_carica"][data-baseweb="button"][kind="secondary"],
    button[key="btn_calcolo"][data-baseweb="button"][kind="secondary"],
    button[key="btn_risultati"][data-baseweb="button"][kind="secondary"],
    button[key="btn_scarica"][data-baseweb="button"][kind="secondary"] {
        opacity: 0.85;
    }
    
    button[key="btn_carica"][data-baseweb="button"][kind="secondary"]:hover,
    button[key="btn_calcolo"][data-baseweb="button"][kind="secondary"]:hover,
    button[key="btn_risultati"][data-baseweb="button"][kind="secondary"]:hover,
    button[key="btn_scarica"][data-baseweb="button"][kind="secondary"]:hover {
        opacity: 1;
    }
    
    /* Info boxes styling */
    .info-modern {
        padding: 1.5rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
        border-left: 4px solid #2196f3;
        margin: 1rem 0;
    }
    
    .success-modern {
        padding: 1.5rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);
        border-left: 4px solid #4caf50;
        margin: 1rem 0;
    }
    
    .warning-modern {
        padding: 1.5rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #fff3e0 0%, #ffe0b2 100%);
        border-left: 4px solid #ff9800;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

# Credenziali di accesso
VALID_USERNAME = "skypemiao"
VALID_PASSWORD = "jfjdljf3244a?091"

# Inizializza session_state per autenticazione
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

def check_credentials(username: str, password: str) -> bool:
    """Verifica le credenziali di accesso"""
    return username == VALID_USERNAME and password == VALID_PASSWORD

def login_page():
    """Mostra la pagina di login"""
    st.markdown('<div class="login-container">', unsafe_allow_html=True)
    st.markdown('<div class="login-title">Accesso Protetto</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-subtitle">Inserisci le credenziali per accedere all\'applicazione</div>', unsafe_allow_html=True)
    
    with st.form("login_form"):
        username = st.text_input("Username", placeholder="Inserisci username")
        password = st.text_input("Password", type="password", placeholder="Inserisci password")
        submit_button = st.form_submit_button("Accedi", use_container_width=True)
        
        if submit_button:
            if check_credentials(username, password):
                st.session_state.authenticated = True
                st.success("Accesso autorizzato")
                st.rerun()
            else:
                st.error("Username o password non corretti. Riprova.")
    
    st.markdown('<div class="company-name">Scay Bergamo</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# Controllo autenticazione
if not st.session_state.authenticated:
    login_page()
    st.stop()

# Header (mostrato solo se autenticato)
# Pulsante Logout in alto al centro
st.markdown('<div class="logout-button-container"></div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([1, 1, 1])
with col2:
    if st.button("Logout", use_container_width=True, type="secondary"):
        st.session_state.authenticated = False
        st.rerun()

st.markdown('<div class="main-header">Calcolo Piano Lavoro - Multi-Tour Operatour</div>', unsafe_allow_html=True)


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalizza i nomi delle colonne"""
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def find_col(df: pd.DataFrame, patterns: List[str]) -> Optional[str]:
    """Trova una colonna che corrisponde a uno dei pattern"""
    cols = [str(c).upper() for c in df.columns]
    for pat in patterns:
        rx = re.compile(pat, re.IGNORECASE)
        for c in cols:
            if rx.search(c):
                return c
    return None


def detect_tour_operators(file_path: str) -> Tuple[Set[str], Set[str]]:
    """
    Rileva tutti i tour operatour unici dal file Excel
    Returns: (tour_operators, aliservice_managed_tour_operators)
    - tour_operators: tutti i tour operator rilevati dalla colonna TOUR OPERATOR
    - aliservice_managed_tour_operators: tour operator gestiti da Aliservice (hanno AGENZIA=Aliservice)
    """
    tour_operators = set()
    aliservice_managed = set()
    
    try:
        xls = pd.ExcelFile(file_path)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            if df is None or df.empty:
                continue
            
            df = normalize_cols(df)
            
            # Rileva TOUR OPERATOR
            to_col = find_col(df, [r"TOUR\s*OPERATOR", r"^TO$", r"OPERATORE"])
            if to_col:
                unique_values = df[to_col].dropna().astype(str).str.strip()
                unique_values = unique_values[unique_values != ""]
                unique_values = unique_values[unique_values.str.lower() != "nan"]
                unique_values = unique_values[unique_values.str.lower() != "none"]
                tour_operators.update(unique_values.unique())
            
            # Rileva AGENZIA per Aliservice e i tour operator che gestisce
            agenzia_col = find_col(df, [r"^AGENZIA$", r"\bAGENCY\b"])
            if agenzia_col and to_col:
                # Trova righe dove AGENZIA contiene "Aliservice"
                mask_aliservice = df[agenzia_col].astype(str).str.contains(r"aliservice", case=False, na=False)
                if mask_aliservice.any():
                    # Aggiungi i tour operator gestiti da Aliservice
                    aliservice_rows = df[mask_aliservice]
                    if to_col in aliservice_rows.columns:
                        managed_tos = aliservice_rows[to_col].dropna().astype(str).str.strip()
                        managed_tos = managed_tos[managed_tos != ""]
                        managed_tos = managed_tos[managed_tos.str.lower() != "nan"]
                        managed_tos = managed_tos[managed_tos.str.lower() != "none"]
                        aliservice_managed.update(managed_tos.unique())
    except Exception as e:
        st.warning(f"Errore nel rilevare tour operatour: {str(e)}")
    
    return tour_operators, aliservice_managed


def find_tour_operator_folder(to_name: str, base_path: str = ".") -> Optional[str]:
    """
    Cerca la cartella del tour operatour nella root e verifica che contenga un file consuntivo*.py
    Returns: path della cartella se trovata e contiene consuntivo*.py, None altrimenti
    """
    to_clean = re.sub(r'[^a-zA-Z]', '', to_name).lower()
    
    if os.path.exists(base_path):
        for item in os.listdir(base_path):
            item_path = os.path.join(base_path, item)
            if os.path.isdir(item_path):
                item_clean = re.sub(r'[^a-zA-Z]', '', item).lower()
                if item_clean == to_clean or to_clean in item_clean or item_clean in to_clean:
                    # Verifica che la cartella contenga un file consuntivo*.py
                    if os.path.exists(item_path):
                        for file in os.listdir(item_path):
                            if file.startswith('consuntivo') and file.endswith('.py'):
                                return item_path
    
    return None


def get_tour_operator_module_name(to_name: str) -> Optional[str]:
    """
    Restituisce il nome normalizzato del tour operator per il matching con i moduli
    Gestisce casi speciali come Baobab/TH, MichelTours, ecc.
    """
    to_clean = re.sub(r'[^a-zA-Z]', '', to_name).lower()
    
    # Casi speciali
    if 'baobab' in to_clean or to_clean == 'th':
        return 'baobab'
    elif 'micheltours' in to_clean or 'michel tours' in to_clean:
        return 'micheltours'
    elif 'aliservice' in to_clean:
        return 'aliservice'
    elif 'caboverdetime' in to_clean:
        return 'caboverdetime'
    elif 'sand' in to_clean:
        return 'sand'
    elif 'rusconi' in to_clean:
        return 'rusconi'
    elif 'domina' in to_clean:
        return 'domina'
    elif 'alpitour' in to_clean:
        return 'alpitour'
    elif 'veratour' in to_clean:
        return 'veratour'
    
    return to_clean


# Valori di default per le opzioni di calcolo (precedentemente nella sidebar)
apt_filter = []  # Nessun filtro per default
night_mode = "DIFF5"  # Default per Veratour
round_extra_mode = "NONE"  # Nessun arrotondamento per default
round_extra_step = 5
round_night_mode = "NONE"  # Nessun arrotondamento per default
round_night_step = 5
holiday_file = None  # Nessun file festivi per default

# Inizializza sezione attiva
if 'active_section' not in st.session_state:
    st.session_state.active_section = 'carica_file'

# Card di navigazione funzionali
col1, col2, col3, col4 = st.columns(4)

with col1:
    is_active = st.session_state.active_section == 'carica_file'
    btn_style = "box-shadow: 0 12px 20px rgba(0, 0, 0, 0.2); transform: translateY(-4px);" if is_active else ""
    if st.button(
        "📁\n\n**Carica File Excel**\n\nCarica il file Excel del piano di lavoro",
        key="btn_carica",
        use_container_width=True,
        type="primary" if is_active else "secondary"
    ):
        st.session_state.active_section = 'carica_file'
        st.rerun()

with col2:
    is_active = st.session_state.active_section == 'esegui_calcolo'
    if st.button(
        "⚙️\n\n**Esegui Calcolo**\n\nAvvia l'elaborazione dei dati",
        key="btn_calcolo",
        use_container_width=True,
        type="primary" if is_active else "secondary"
    ):
        st.session_state.active_section = 'esegui_calcolo'
        st.rerun()

with col3:
    is_active = st.session_state.active_section == 'risultati'
    if st.button(
        "📊\n\n**Risultati**\n\nVisualizza i risultati dell'elaborazione",
        key="btn_risultati",
        use_container_width=True,
        type="primary" if is_active else "secondary"
    ):
        st.session_state.active_section = 'risultati'
        st.rerun()

with col4:
    is_active = st.session_state.active_section == 'scarica_risultati'
    if st.button(
        "💾\n\n**Scarica Risultati**\n\nScarica il file Excel generato",
        key="btn_scarica",
        use_container_width=True,
        type="primary" if is_active else "secondary"
    ):
        st.session_state.active_section = 'scarica_risultati'
        st.rerun()

# Area principale - Mostra solo la sezione attiva
st.markdown("---")

# Sezione Carica File Excel
if st.session_state.active_section == 'carica_file':
    st.markdown("### 📁 Carica File Excel del Piano di Lavoro")
    uploaded_file = st.file_uploader(
        "Seleziona il file Excel",
        type=["xlsx", "xls"],
        help="Carica il file Excel contenente il piano di lavoro"
    )
    if uploaded_file is not None:
        st.session_state['uploaded_file'] = uploaded_file
        # Processa immediatamente il file caricato
        uploaded_file = st.session_state['uploaded_file']
    else:
        uploaded_file = None
else:
    uploaded_file = st.session_state.get('uploaded_file', None)

# Mostra info file solo nella sezione carica_file
if st.session_state.active_section == 'carica_file' and uploaded_file is not None:
    file_details = {
        "Nome file": uploaded_file.name,
        "Dimensione": f"{uploaded_file.size / 1024:.2f} KB",
        "Tipo": uploaded_file.type
    }
    
    with st.expander("ℹ️ Informazioni File", expanded=False):
        st.json(file_details)
    
    # Rileva tour operatour
    with st.spinner("🔍 Rilevamento tour operatour..."):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_path = tmp_file.name
        
        try:
            tour_operators, aliservice_managed = detect_tour_operators(tmp_path)
            
            # Verifica se Aliservice è disponibile (se ci sono tour operator gestiti da Aliservice)
            aliservice_available = False
            if aliservice_managed:
                # Verifica se esiste la cartella Aliservice
                aliservice_folder = find_tour_operator_folder("Aliservice")
                if aliservice_folder:
                    aliservice_available = True
            
            # Crea lista completa: tutti i tour operator rilevati + Aliservice se disponibile
            all_tour_operators = set(tour_operators)
            if aliservice_available:
                all_tour_operators.add("ALISERVICE")
            
            if all_tour_operators:
                st.info(f"Tour operatour rilevati: {', '.join(sorted(all_tour_operators))}")
                
                available_folders = {}
                missing_tour_operators = []
                
                # Per ogni tour operator, verifica se esiste la cartella di calcolo
                for to_name in sorted(all_tour_operators):
                    # Skip i tour operator gestiti da Aliservice (sono gestiti da Aliservice)
                    if to_name in aliservice_managed and aliservice_available:
                        continue
                    
                    folder_path = find_tour_operator_folder(to_name)
                    if folder_path:
                        available_folders[to_name] = folder_path
                    else:
                        missing_tour_operators.append(to_name)
                
            if available_folders:
                st.success(f"Tour operatour con calcolo disponibile: {', '.join(sorted(available_folders.keys()))}")
            
            if missing_tour_operators:
                st.warning(f"Tour operatour senza cartella (non elaborati): {', '.join(sorted(missing_tour_operators))}")
        finally:
            pass  # Manteniamo tmp_path per il calcolo
    
    # Salva tmp_path in session state per il calcolo
    st.session_state['tmp_file_path'] = tmp_path
    st.session_state['uploaded_file_name'] = uploaded_file.name

# Sezione Esegui Calcolo
if st.session_state.active_section == 'esegui_calcolo':
    st.markdown("### ⚙️ Esegui Calcolo")
    uploaded_file = st.session_state.get('uploaded_file', None)
    
    if uploaded_file is None:
        st.warning("⚠️ Nessun file caricato. Vai alla sezione 'Carica File Excel' per caricare un file.")
    else:
        if st.button("Esegui Calcolo", type="primary", use_container_width=True):
            with st.spinner("⏳ Elaborazione in corso..."):
                try:
                    # Usa il file temporaneo già creato durante il caricamento
                    tmp_path = st.session_state.get('tmp_file_path')
                    if not tmp_path or not os.path.exists(tmp_path):
                        # Se non esiste, ricrearlo
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                            tmp_file.write(uploaded_file.getvalue())
                            tmp_path = tmp_file.name
                    
                    try:
                        # Carica festivi se presente
                        holiday_dates = None
                        if holiday_file is not None:
                            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix=".txt") as tmp_holiday:
                                tmp_holiday.write(holiday_file.getvalue().decode('utf-8'))
                                tmp_holiday_path = tmp_holiday.name
                            try:
                                holiday_dates = load_holiday_list(tmp_holiday_path)
                            finally:
                                os.unlink(tmp_holiday_path)
                        
                        # Rileva tour operatour
                        tour_operators, aliservice_managed = detect_tour_operators(tmp_path)
                        
                        # Verifica se Aliservice è presente (se ci sono tour operator gestiti da Aliservice)
                        aliservice_found = False
                        aliservice_folder = find_tour_operator_folder("Aliservice")
                        if aliservice_managed and aliservice_folder:
                            # Verifica se ALISERVICE_AVAILABLE è definito, altrimenti assume False
                            aliservice_available = globals().get('ALISERVICE_AVAILABLE', False)
                            if aliservice_available:
                                aliservice_found = True
                        
                        # Rileva dinamicamente quali tour operatour hanno la cartella di calcolo
                        # Filtra i tour operator gestiti da Aliservice dalla lista principale
                        tour_operators_to_check = tour_operators - aliservice_managed
                        
                        # Mappa i tour operator trovati ai loro nomi normalizzati per il matching
                        found_tour_operators = {}
                        for to_name in tour_operators_to_check:
                            folder_path = find_tour_operator_folder(to_name)
                            if folder_path:
                                module_name = get_tour_operator_module_name(to_name)
                                if module_name:
                                    found_tour_operators[module_name] = {
                                        'original_name': to_name,
                                        'folder': folder_path
                                    }
                        
                        all_detail_dfs = []
                        all_totals_dfs = []
                        all_discr_dfs = []
                        
                        # Dizionario di mapping tra nomi normalizzati e funzioni/classi disponibili
                        tour_operator_processors = {
                        'veratour': {
                            'available': True,  # Veratour è sempre disponibile
                            'config_class': VeratourCalcConfig,
                            'process_func': process_files_veratour,
                            'write_func': write_output_excel_veratour,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'night_mode': night_mode,
                                'rounding_extra': RoundingPolicy(round_extra_mode, round_extra_step),
                                'rounding_night': RoundingPolicy(round_night_mode, round_night_step),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'alpitour': {
                            'available': ALPITOUR_AVAILABLE,
                            'config_class': AlpitourCalcConfig,
                            'process_func': process_files_alpitour,
                            'write_func': write_output_excel_alpitour,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'aliservice': {
                            'available': globals().get('ALISERVICE_AVAILABLE', False),
                            'config_class': AliserviceCalcConfig,
                            'process_func': process_files_aliservice,
                            'write_func': write_output_excel_aliservice,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'baobab': {
                            'available': globals().get('BAOBAB_AVAILABLE', False),
                            'config_class': BaobabCalcConfig,
                            'process_func': process_files_baobab,
                            'write_func': write_output_excel_baobab,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'domina': {
                            'available': globals().get('DOMINA_AVAILABLE', False),
                            'config_class': DominaCalcConfig,
                            'process_func': process_files_domina,
                            'write_func': write_output_excel_domina,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'micheltours': {
                            'available': globals().get('MICHELTOURS_AVAILABLE', False),
                            'config_class': MichelToursCalcConfig,
                            'process_func': process_files_micheltours,
                            'write_func': write_output_excel_micheltours,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'sand': {
                            'available': globals().get('SAND_AVAILABLE', False),
                            'config_class': SandCalcConfig,
                            'process_func': process_files_sand,
                            'write_func': write_output_excel_sand,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'caboverdetime': {
                            'available': globals().get('CABOVERDETIME_AVAILABLE', False),
                            'config_class': CaboverdetimeCalcConfig,
                            'process_func': process_files_caboverdetime,
                            'write_func': write_output_excel_caboverdetime,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                        'rusconi': {
                            'available': globals().get('RUSCONI_AVAILABLE', False),
                            'config_class': RusconiCalcConfig,
                            'process_func': process_files_rusconi,
                            'write_func': write_output_excel_rusconi,
                            'config_kwargs': lambda: {
                                'apt_filter': apt_filter if apt_filter else None,
                                'rounding_extra': RoundingPolicy("NONE", 5),
                                'rounding_night': RoundingPolicy("NONE", 5),
                                'holiday_dates': holiday_dates,
                            }
                        },
                    }
                    
                        # Raccogli prima tutti i tour operator da elaborare per mostrarli in orizzontale
                        tour_operators_to_process = []
                        
                        # Prepara lista tour operator normali
                        warnings_list = []
                        for module_name, to_info in found_tour_operators.items():
                            if module_name in tour_operator_processors:
                                processor = tour_operator_processors[module_name]
                                if processor['available'] and processor['config_class'] and processor['process_func']:
                                    tour_operators_to_process.append({
                                        'name': to_info['original_name'],
                                        'module_name': module_name,
                                        'processor': processor,
                                        'is_aliservice': False
                                    })
                                else:
                                    warnings_list.append(f"{to_info['original_name']} rilevato ma modulo non disponibile. Installare le dipendenze necessarie.")
                        
                        # Aggiungi Aliservice se presente
                        if aliservice_found and 'aliservice' in tour_operator_processors:
                            processor = tour_operator_processors['aliservice']
                            if processor['available'] and processor['config_class'] and processor['process_func']:
                                tour_operators_to_process.append({
                                    'name': 'Aliservice',
                                    'module_name': 'aliservice',
                                    'processor': processor,
                                    'is_aliservice': True
                                })
                            else:
                                warnings_list.append("Aliservice rilevato ma modulo non disponibile. Installare le dipendenze necessarie.")
                        
                        # Mostra eventuali warning
                        if warnings_list:
                            for warning in warnings_list:
                                st.warning(warning)
                        
                        # Mostra i tour operator che verranno elaborati in colonne orizzontali
                        if tour_operators_to_process:
                            # Calcola il numero di colonne (massimo 4 per riga)
                            num_cols = min(len(tour_operators_to_process), 4)
                            cols = st.columns(num_cols)
                            for idx, to_info in enumerate(tour_operators_to_process):
                                col_idx = idx % num_cols
                                with cols[col_idx]:
                                    st.info(f"Elaborazione {to_info['name']}...")
                        
                        # Elabora dinamicamente tutti i tour operator
                        processed_count = 0
                        errors = []
                        
                        for to_info in tour_operators_to_process:
                            processor = to_info['processor']
                            try:
                                cfg = processor['config_class'](**processor['config_kwargs']())
                                detail, totals, discr = processor['process_func']([tmp_path], cfg)
                                all_detail_dfs.append(detail)
                                all_totals_dfs.append(totals)
                                all_discr_dfs.append(discr)
                                processed_count += 1
                            except Exception as e:
                                errors.append(f"{to_info['name']}: {str(e)}")
                        
                        # Mostra eventuali errori
                        if errors:
                            for error in errors:
                                st.error(f"Errore durante l'elaborazione: {error}")
                        
                        if processed_count == 0:
                            st.error("Nessun tour operatour con calcolo disponibile trovato nel file. Verifica che le cartelle dei tour operator siano presenti e contengano i file consuntivo*.py.")
                            st.stop()
                        
                        # Combina i risultati
                        if all_detail_dfs:
                            detail_df = pd.concat(all_detail_dfs, ignore_index=True)
                            totals_df = pd.concat(all_totals_dfs, ignore_index=True)
                            # Combina discrepanze (filtra quelli non vuoti)
                            discr_list = [d for d in all_discr_dfs if d is not None and not d.empty]
                            if discr_list:
                                discr_df = pd.concat(discr_list, ignore_index=True)
                            else:
                                discr_df = pd.DataFrame()
                        else:
                            detail_df = pd.DataFrame()
                            totals_df = pd.DataFrame()
                            discr_df = pd.DataFrame()
                        
                        # Genera output Excel
                        output_buffer = io.BytesIO()
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:
                            output_path = tmp_output.name
                        
                            # Usa la funzione di scrittura appropriata
                            # Se ci sono più tour operator, usa Alpitour (supporta tutti i fogli e gestisce meglio i dati combinati)
                            # Altrimenti usa la funzione specifica del primo tour operator processato
                            if processed_count > 1 and ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                                # Più tour operator: usa Alpitour (supporta tutti i fogli e gestisce meglio i dati combinati)
                                write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
                            elif processed_count > 0:
                                # Usa la funzione di scrittura del primo tour operator processato
                                first_processed = list(found_tour_operators.keys())[0] if found_tour_operators else None
                                if aliservice_found and 'aliservice' in tour_operator_processors:
                                    processor = tour_operator_processors['aliservice']
                                    if processor.get('write_func'):
                                        processor['write_func'](output_path, detail_df, totals_df, discr_df)
                                    elif ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                                        write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
                                    else:
                                        write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
                                elif first_processed and first_processed in tour_operator_processors:
                                    processor = tour_operator_processors[first_processed]
                                    if processor.get('write_func'):
                                        processor['write_func'](output_path, detail_df, totals_df, discr_df)
                                    elif ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                                        write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
                                    else:
                                        write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
                                else:
                                    # Fallback a Alpitour o Veratour
                                    if ALPITOUR_AVAILABLE and write_output_excel_alpitour:
                                        write_output_excel_alpitour(output_path, detail_df, totals_df, discr_df)
                                    else:
                                        write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
                            else:
                                # Fallback a Veratour se nessun altro disponibile
                                write_output_excel_veratour(output_path, detail_df, totals_df, discr_df)
                    
                            # Aggiungi foglio TourOperatourRilevati con tutti i tour operator (codificati e non)
                            from openpyxl import load_workbook
                            from openpyxl.styles import Font, PatternFill
                            wb = load_workbook(output_path)
                            
                            # Rimuovi foglio esistente se presente
                            if "TourOperatourRilevati" in wb.sheetnames:
                                wb.remove(wb["TourOperatourRilevati"])
                            if "TourOperatourNonElaborati" in wb.sheetnames:
                                wb.remove(wb["TourOperatourNonElaborati"])
                            
                            ws = wb.create_sheet("TourOperatourRilevati", 0)  # Crea come primo foglio
                            
                            # Raccogli i tour operator effettivamente elaborati dal detail_df
                            elaborated_tour_operators = set()
                            if not detail_df.empty and 'TOUR OPERATOR' in detail_df.columns:
                                elaborated_tour_operators = set(detail_df['TOUR OPERATOR'].dropna().astype(str).unique())
                            
                            # Crea lista di tutti i tour operator con il loro status
                            # Aggiungi Aliservice se presente (è un'agenzia, non un tour operator nella lista principale)
                            tour_operators_for_list = tour_operators - aliservice_managed
                            if aliservice_found:
                                tour_operators_for_list.add("ALISERVICE")
                            
                            tour_operator_list = []
                            for to_name in sorted(tour_operators_for_list):
                                to_clean = re.sub(r'[^a-zA-Z]', '', to_name).lower()
                                is_supported = False
                                status = "Non codificato"
                                
                                # Caso speciale per Aliservice (può avere anche AGENZIA nel detail_df)
                                if to_name.upper() == "ALISERVICE" and aliservice_found:
                                    # Verifica se Aliservice appare nel detail_df (potrebbe essere in AGENZIA)
                                    if not detail_df.empty:
                                        # Controlla se c'è una colonna AGENZIA con Aliservice
                                        if 'AGENZIA' in detail_df.columns:
                                            if detail_df['AGENZIA'].astype(str).str.contains('aliservice', case=False, na=False).any():
                                                is_supported = True
                                                status = "Codificato - Elaborato"
                                        # Oppure controlla TOUR OPERATOR se contiene dati di Aliservice
                                        elif 'TOUR OPERATOR' in detail_df.columns:
                                            for elaborated_to in elaborated_tour_operators:
                                                if 'aliservice' in str(elaborated_to).lower():
                                                    is_supported = True
                                                    status = "Codificato - Elaborato"
                                                    break
                                    
                                    if not is_supported:
                                        is_supported = True
                                        status = "Codificato - Rilevato ma senza dati elaborati"
                                else:
                                    # Verifica se è stato effettivamente elaborato (appare nel detail_df)
                                    if not detail_df.empty and 'TOUR OPERATOR' in detail_df.columns:
                                        # Confronta il nome normalizzato
                                        for elaborated_to in elaborated_tour_operators:
                                            elaborated_clean = re.sub(r'[^a-zA-Z]', '', str(elaborated_to)).lower()
                                            if to_clean == elaborated_clean or to_clean in elaborated_clean or elaborated_clean in to_clean:
                                                is_supported = True
                                                status = "Codificato - Elaborato"
                                                break
                                    
                                    # Se non trovato nel detail_df, verifica se è tra quelli supportati trovati
                                    if not is_supported:
                                        # Verifica dinamicamente se il tour operator ha una cartella di calcolo
                                        module_name = get_tour_operator_module_name(to_name)
                                        if module_name and module_name in found_tour_operators:
                                            is_supported = True
                                            status = "Codificato - Rilevato ma senza dati elaborati"
                                        elif module_name and module_name in tour_operator_processors:
                                            # Ha un processore ma non è stato trovato nel file (potrebbe essere un problema di matching)
                                            processor = tour_operator_processors[module_name]
                                            if processor['available'] and processor['config_class']:
                                                is_supported = True
                                                status = "Codificato - Rilevato ma senza dati elaborati"
                                
                                # Se ancora non supportato, verifica se ha una cartella (potrebbe essere codificato ma non rilevato)
                                if not is_supported:
                                    folder_path = find_tour_operator_folder(to_name)
                                    if folder_path:
                                        status = "Modulo presente ma non rilevato nel file"
                                    else:
                                        status = "Non codificato"
                                
                                note = "Calcolo tariffe disponibile e applicato" if status == "Codificato - Elaborato" else \
                                       "Calcolo tariffe disponibile ma nessun dato da elaborare" if "Codificato" in status else \
                                       "Calcolo tariffe non disponibile - da codificare"
                                
                                tour_operator_list.append({
                                    "Tour Operatour": to_name,
                                    "Status": status,
                                    "Note": note
                                })
                            
                            # Scrivi header
                            ws.cell(row=1, column=1, value="Tour Operatour")
                            ws.cell(row=1, column=2, value="Status")
                            ws.cell(row=1, column=3, value="Note")
                            
                            # Scrivi dati
                            for idx, to_info in enumerate(tour_operator_list, 2):
                                ws.cell(row=idx, column=1, value=to_info["Tour Operatour"])
                                ws.cell(row=idx, column=2, value=to_info["Status"])
                                ws.cell(row=idx, column=3, value=to_info["Note"])
                            
                            # Formatta header in grassetto
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF")
                            for col in range(1, 4):
                                cell = ws.cell(row=1, column=col)
                                cell.fill = header_fill
                                cell.font = header_font
                            
                            # Auto-adjust column widths
                            ws.column_dimensions['A'].width = max(30, len("Tour Operatour") + 5)
                            ws.column_dimensions['B'].width = max(30, len("Status") + 5)
                            ws.column_dimensions['C'].width = max(60, len("Note") + 5)
                            
                            wb.save(output_path)
                            
                            # Leggi il file generato
                            with open(output_path, 'rb') as f:
                                output_buffer.write(f.read())
                            output_buffer.seek(0)
                            
                            # Salva in session state
                            st.session_state['output_file'] = output_buffer
                            st.session_state['output_filename'] = f"OUT_{uploaded_file.name}"
                            st.session_state['detail_df'] = detail_df
                            st.session_state['totals_df'] = totals_df
                            st.session_state['discr_df'] = discr_df
                        
                        # Cleanup
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                        if os.path.exists(output_path):
                            os.unlink(output_path)
                        # Rimuovi da session state
                        if 'tmp_file_path' in st.session_state:
                            del st.session_state['tmp_file_path']
                        
                        st.success(f"Calcolo completato! Blocchi calcolati: {len(detail_df)}")
                        
                        if not discr_df.empty:
                            st.warning(f"Discrepanze rilevate: {len(discr_df)} (vedi sezione Discrepanze)")
                    
                    except Exception as e:
                        st.error(f"Errore durante l'elaborazione: {str(e)}")
                        st.exception(e)
                        if 'tmp_file_path' in st.session_state and os.path.exists(st.session_state['tmp_file_path']):
                            os.unlink(st.session_state['tmp_file_path'])
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                
                except Exception as e:
                    st.error(f"Errore: {str(e)}")
                    st.exception(e)

# Sezione Risultati
elif st.session_state.active_section == 'risultati':
    st.markdown("### 📊 Risultati")
    
    if 'output_file' in st.session_state and st.session_state['output_file'] is not None:
        if 'discr_df' in st.session_state and not st.session_state['discr_df'].empty:
            with st.expander("Discrepanze Rilevate", expanded=True):
                st.dataframe(st.session_state['discr_df'], use_container_width=True)
                st.info("Le discrepanze indicano possibili inconsistenze nei dati di input. Controlla il file Excel generato per i dettagli completi.")
        else:
            st.success("✅ Nessuna discrepanza rilevata. Il calcolo è stato completato con successo!")
    else:
        st.info("⚠️ Nessun risultato disponibile. Esegui prima il calcolo nella sezione 'Esegui Calcolo'.")

# Sezione Scarica Risultati
elif st.session_state.active_section == 'scarica_risultati':
    st.markdown("### 💾 Scarica Risultati")
    
    if 'output_file' in st.session_state and st.session_state['output_file'] is not None:
        st.download_button(
            label="📥 Scarica File Excel Completo",
            data=st.session_state['output_file'],
            file_name=st.session_state['output_filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.info("""
        Il file generato conterrà:
        - Fogli dettagliati per ogni aeroporto (VRN, BGY, NAP, VCE)
        - Foglio TOTALE con i riepiloghi
        - Fogli tecnici (DettaglioBlocchi, TotaliPeriodo, Discrepanze)
        - Foglio TourOperatourRilevati con lo stato di tutti i tour operator
        """)
    else:
        st.warning("⚠️ Nessun file disponibile per il download. Esegui prima il calcolo nella sezione 'Esegui Calcolo'.")

# Messaggio iniziale se nessuna sezione è attiva o se non c'è file caricato
if st.session_state.active_section == 'carica_file' and uploaded_file is None:
    st.info("""
    👋 **Benvenuto nel Calcolatore Piano Lavoro Multi-Tour Operatour!**
    
    Per iniziare:
    1. Carica il file Excel del piano di lavoro usando la card "Carica File Excel" sopra
    2. Clicca su "Esegui Calcolo" per avviare l'elaborazione
    3. Visualizza i "Risultati" e "Scarica Risultati" quando disponibili
    
    Il file generato conterrà:
    - Fogli dettagliati per ogni aeroporto (VRN, BGY, NAP, VCE)
    - Foglio TOTALE con i riepiloghi
    - Fogli tecnici (DettaglioBlocchi, TotaliPeriodo, Discrepanze)
    - Foglio TourOperatourRilevati con lo stato di tutti i tour operator
    """)

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666; padding: 1rem;'>"
    "Calcolo Piano Lavoro - Multi-Tour Operatour | Scay"
    "</div>",
    unsafe_allow_html=True
)
