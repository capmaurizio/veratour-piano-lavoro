#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analizza il template di riepilogo assistenze"""

import openpyxl

file_path = "Documentazione/RIEPILOGO ASSISTENZE  BGY SORRENTI GAIA dicembre-febbraio 25 .xlsx"

wb = openpyxl.load_workbook(file_path, data_only=False)
print("Fogli:", wb.sheetnames)
ws = wb.active
print(f"Foglio attivo: {ws.title}")
print(f"Righe: {ws.max_row}, Colonne: {ws.max_column}")

print("\n=== Prime 15 righe ===")
for i, row in enumerate(ws.iter_rows(max_row=15, values_only=False), 1):
    values = []
    for cell in row[:15]:
        if cell.value:
            values.append(str(cell.value)[:30])
        else:
            values.append("")
    print(f"Riga {i}: {' | '.join(values)}")

print("\n=== Formule trovate ===")
formule_trovate = []
for row in ws.iter_rows():
    for cell in row:
        if cell.data_type == 'f':
            formule_trovate.append((cell.coordinate, cell.value))
            print(f"{cell.coordinate}: {cell.value}")

print(f"\nTotale formule: {len(formule_trovate)}")

# Cerca riferimenti al nome assistente
print("\n=== Cerca riferimenti a 'GAIA' o 'SORRENTI' ===")
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if 'GAIA' in str(cell.value).upper() or 'SORRENTI' in str(cell.value).upper():
                print(f"{cell.coordinate}: {cell.value}")
