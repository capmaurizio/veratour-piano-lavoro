#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera un template Excel personalizzato per ogni assistente
basato sul file RIEPILOGO ASSISTENZE
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import os
import tempfile
import pandas as pd
from typing import Optional, Dict
import sys
import importlib.util

# Path del template (relativo alla directory del progetto)
TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "Documentazione",
    "RIEPILOGO ASSISTENZE  BGY SORRENTI GAIA .xlsx"
)

# Tariffa extra oraria per aeroporto (Accordo 2026) - usata per le formule Excel
# NAP Senior: €12/h, NAP Junior: €10/h, FCO: €12/h, FCO Incentive: €15/h, VRN: €12/h, BGY Senior: €10/h, BGY Junior: €8/h
TARIFFE_EXTRA_PER_APT = {
    'NAP': 12.0,   # Senior default — per junior cambia in 10.0
    'FCO': 12.0,
    'VRN': 12.0,
    'BGY': 8.0,    # Junior default — per senior cambia in 10.0
    'MXP': 12.0,
    'CTA': 12.0,
    'PMO': 12.0,
    'PSA': 12.0,
    'BRI': 12.0,
    'BLQ': 12.0,
    'VCE': 12.0,
    'TSF': 12.0,
}

def parse_time_value(val) -> Optional[time]:
    """Converte un valore in time object"""
    if pd.isna(val):
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ['nan', 'none', '']:
        return None
    try:
        parts = val_str.split(':')
        if len(parts) >= 2:
            h = int(parts[0])
            m = int(parts[1])
            return time(h, m)
    except:
        pass
    return None

def parse_date_value(val) -> Optional[date]:
    """Converte un valore in date object"""
    if pd.isna(val):
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, pd.Timestamp):
        return val.date()
    return None

def carica_file_calcolo_assistente(nome_assistente: str):
    """
    Carica il file Python di calcolo per un assistente
    
    Args:
        nome_assistente: Nome dell'assistente
    
    Returns:
        Modulo Python caricato o None
    """
    nome_safe = nome_assistente.replace(" ", "_").replace("/", "_").upper()
    file_path = os.path.join(os.path.dirname(__file__), "calcoli_assistenti", f"calcolo_{nome_safe}.py")
    
    if not os.path.exists(file_path):
        return None
    
    try:
        spec = importlib.util.spec_from_file_location(f"calcolo_{nome_safe}", file_path)
        modulo = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(modulo)
        return modulo
    except Exception as e:
        print(f"Errore nel caricare file calcolo per {nome_assistente}: {e}")
        return None

def genera_template_assistente(
    nome_assistente: str, 
    output_path: str = None,
    turni_piano_lavoro: Optional[pd.DataFrame] = None,
    dati_salvati: Optional[Dict] = None
) -> str:
    """
    Genera un template Excel personalizzato per un assistente
    con i dati del piano lavoro già compilati
    
    Args:
        nome_assistente: Nome dell'assistente
        output_path: Path dove salvare il file (opzionale)
        turni_piano_lavoro: DataFrame con i turni del piano lavoro per questo assistente
        dati_salvati: Dizionario con i dati salvati dell'assistente (ore effettive, extra, notte)
    
    Returns:
        Path del file generato
    """
    
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template non trovato: {TEMPLATE_PATH}")
    
    # Carica il template originale
    wb_template = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    
    # Crea nuovo workbook
    wb_nuovo = openpyxl.Workbook()
    
    # Rimuovi foglio di default
    if 'Sheet' in wb_nuovo.sheetnames:
        wb_nuovo.remove(wb_nuovo['Sheet'])
    
    # Copia tutti i fogli dal template
    for nome_foglio in wb_template.sheetnames:
        ws_template = wb_template[nome_foglio]
        ws_nuovo = wb_nuovo.create_sheet(title=nome_foglio)
        
        # Copia tutte le celle preservando formule e stili
        # Per il foglio "DICEMBRE 2025", non copiare i dati di esempio, solo stili e formule
        is_dicembre_2025 = ('DICEMBRE 2025' in nome_foglio.upper() or 'DICEMBRE' in nome_foglio.upper())
        riga_header = 1  # Assumiamo che la riga 1 sia l'header
        
        for row in ws_template.iter_rows():
            for cella_template in row:
                if cella_template.value is not None or cella_template.has_style:
                    cella_nuova = ws_nuovo.cell(
                        row=cella_template.row,
                        column=cella_template.column
                    )
                    
                    # Copia valore o formula
                    if cella_template.data_type == 'f':  # Formula
                        # Preserva sempre le formule
                        cella_nuova.value = cella_template.value
                    else:
                        # Per il foglio "DICEMBRE 2025", copia solo header (riga 1), non i dati di esempio
                        if is_dicembre_2025 and cella_template.row > riga_header:
                            # Non copiare i dati di esempio, solo gli stili (ma senza sfondo nero)
                            cella_nuova.value = None
                        else:
                            # Per altri fogli o header, copia tutto
                            cella_nuova.value = cella_template.value
                    
                    # Copia stili
                    try:
                        # Fill - NON copiare se è nero o se è una riga dati nel foglio DICEMBRE 2025
                        if cella_template.fill and cella_template.fill.start_color:
                            rgb = cella_template.fill.start_color.rgb
                            if rgb:
                                # Verifica se è nero (000000 o simile)
                                rgb_str = str(rgb).upper()
                                is_nero = rgb_str in ['000000', '00000000', 'FF000000', 'FF00000000'] or '000000' in rgb_str
                                
                                # Se è una riga dati nel foglio DICEMBRE 2025 e lo sfondo è nero, non copiare
                                if is_dicembre_2025 and cella_template.row > riga_header and is_nero:
                                    # Non applicare sfondo nero alle righe dati
                                    pass
                                else:
                                    # Copia lo sfondo normale
                                    cella_nuova.fill = PatternFill(
                                        start_color=cella_template.fill.start_color.rgb,
                                        end_color=cella_template.fill.end_color.rgb if cella_template.fill.end_color and cella_template.fill.end_color.rgb else cella_template.fill.start_color.rgb,
                                        fill_type=cella_template.fill.fill_type or 'solid'
                                    )
                        
                        # Font
                        if cella_template.font:
                            font_color = None
                            if cella_template.font.color and hasattr(cella_template.font.color, 'rgb'):
                                font_color = cella_template.font.color.rgb
                            
                            cella_nuova.font = Font(
                                name=cella_template.font.name,
                                size=cella_template.font.size,
                                bold=cella_template.font.bold,
                                italic=cella_template.font.italic,
                                color=font_color
                            )
                        
                        # Alignment
                        if cella_template.alignment:
                            cella_nuova.alignment = Alignment(
                                horizontal=cella_template.alignment.horizontal,
                                vertical=cella_template.alignment.vertical,
                                wrap_text=cella_template.alignment.wrap_text,
                                shrink_to_fit=cella_template.alignment.shrink_to_fit,
                                indent=cella_template.alignment.indent
                            )
                        
                        # Border
                        if cella_template.border:
                            border_template = cella_template.border
                            cella_nuova.border = Border(
                                left=border_template.left,
                                right=border_template.right,
                                top=border_template.top,
                                bottom=border_template.bottom,
                                diagonal=border_template.diagonal,
                                diagonal_direction=border_template.diagonal_direction,
                                outline=border_template.outline,
                                vertical=border_template.vertical,
                                horizontal=border_template.horizontal
                            )
                        
                        # Number format
                        if cella_template.number_format:
                            cella_nuova.number_format = cella_template.number_format
                    except:
                        pass
        
        # Copia larghezza colonne
        for col_letter in ws_template.column_dimensions:
            if col_letter in ws_template.column_dimensions:
                ws_nuovo.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width
        
        # Copia altezza righe
        for row_num in ws_template.row_dimensions:
            if row_num in ws_template.row_dimensions:
                ws_nuovo.row_dimensions[row_num].height = ws_template.row_dimensions[row_num].height
        
        # Copia merged cells
        for merged_range in ws_template.merged_cells.ranges:
            ws_nuovo.merge_cells(str(merged_range))
    
    # Carica file calcolo per questo assistente
    modulo_calcolo = carica_file_calcolo_assistente(nome_assistente)
    
    # Popola il foglio "Dicembre 2025" con i dati del piano lavoro (se disponibili)
    # Cerca il foglio "Dicembre 2025" o usa il primo foglio
    foglio_target = None
    for nome_foglio in wb_nuovo.sheetnames:
        if 'DICEMBRE 2025' in nome_foglio.upper() or 'DICEMBRE' in nome_foglio.upper():
            foglio_target = nome_foglio
            break
    
    if not foglio_target and len(wb_nuovo.sheetnames) > 0:
        foglio_target = wb_nuovo.sheetnames[0]
    
    if turni_piano_lavoro is not None and not turni_piano_lavoro.empty and foglio_target:
        ws_riepilogo = wb_nuovo[foglio_target]
        
        # Trova la riga di inizio dati (dopo l'header, di solito riga 2)
        riga_inizio = 2
        
        # I dati di esempio sono già stati puliti durante la copia (solo per il foglio riepilogo)
        # Non serve pulire di nuovo
        
        # Mappa colonne del piano lavoro
        def trova_colonna(df, patterns):
            for pattern in patterns:
                for col in df.columns:
                    if pattern.upper() in str(col).upper():
                        return col
            return None
        
        # Funzione helper per scrivere in una cella (gestisce merged cells e rimuove sfondi neri)
        def scrivi_cella_sicura(ws, row, col, value):
            """Scrive in una cella gestendo correttamente le celle unite e rimuove sfondi neri"""
            try:
                # Verifica se la cella fa parte di un merge
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        # È parte di un merge, scrivi nella cella principale (top-left)
                        cella_principale = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        cella_principale.value = value
                        # Rimuovi sfondo nero se presente
                        if cella_principale.fill and cella_principale.fill.start_color:
                            rgb = str(cella_principale.fill.start_color.rgb or '').upper()
                            if '000000' in rgb:
                                cella_principale.fill = PatternFill()  # Sfondo vuoto
                        return
                
                # Non è parte di un merge, scrivi normalmente
                cella = ws.cell(row=row, column=col)
                # Verifica se è una MergedCell (read-only) - in tal caso salta
                if hasattr(cella, '__class__') and 'MergedCell' in str(cella.__class__):
                    # È una MergedCell read-only, non possiamo scrivere
                    return
                
                cella.value = value
                # Rimuovi sfondo nero se presente
                if cella.fill and cella.fill.start_color:
                    rgb = str(cella.fill.start_color.rgb or '').upper()
                    if '000000' in rgb:
                        cella.fill = PatternFill()  # Sfondo vuoto (bianco)
            except AttributeError:
                # Se è una MergedCell, prova a trovare la cella principale
                try:
                    for merged_range in ws.merged_cells.ranges:
                        if (merged_range.min_row <= row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                            cella_princ = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            cella_princ.value = value
                            # Rimuovi sfondo nero
                            if cella_princ.fill and cella_princ.fill.start_color:
                                rgb = str(cella_princ.fill.start_color.rgb or '').upper()
                                if '000000' in rgb:
                                    cella_princ.fill = PatternFill()
                            return
                except:
                    pass
            except Exception:
                # Se fallisce, ignora
                pass
        
        # Ordina i turni per data
        turni_ordinati = turni_piano_lavoro.copy()
        # Aggiungi colonna data parsata per ordinare
        def parse_data_for_sort(val):
            data_val = parse_date_value(val)
            return data_val if data_val else datetime(1900, 1, 1).date()
        
        turni_ordinati['_DATA_SORT'] = turni_ordinati.get('DATA', pd.Series()).apply(parse_data_for_sort)
        turni_ordinati = turni_ordinati.sort_values('_DATA_SORT')
        
        # Popola i dati per ogni turno e genera formule
        righe_dati = []  # Traccia le righe con dati per le formule di totale
        
        for idx, (_, row_piano) in enumerate(turni_ordinati.iterrows(), start=riga_inizio):
            # DATA (colonna A)
            data_val = parse_date_value(row_piano.get('DATA', None))
            if data_val:
                scrivi_cella_sicura(ws_riepilogo, idx, 1, data_val)
            
            # Operatore / Tour Operator (colonna B)
            tour_op = str(row_piano.get('TOUR OPERATOR', '')).strip()
            if tour_op:
                scrivi_cella_sicura(ws_riepilogo, idx, 2, tour_op)
            
            # Tipo assistenza / Servizio (colonna C)
            servizio = str(row_piano.get('SERVIZIO', '')).strip()
            if not servizio:
                servizio = str(row_piano.get('SERVIZI', '')).strip()
            if servizio:
                scrivi_cella_sicura(ws_riepilogo, idx, 3, servizio)
            
            # Convoc / STD (colonna D)
            std_val = parse_time_value(row_piano.get('STD', None))
            if std_val:
                scrivi_cella_sicura(ws_riepilogo, idx, 4, std_val.strftime('%H:%M'))
            
            # Volo (colonna E)
            volo = str(row_piano.get('VOLO', '')).strip()
            if volo and volo.lower() not in ['nan', 'none', '']:
                scrivi_cella_sicura(ws_riepilogo, idx, 5, volo)
            
            # Previsto / STD (colonna F)
            if std_val:
                scrivi_cella_sicura(ws_riepilogo, idx, 6, std_val.strftime('%H:%M'))
            
            # Effettivo / ATD (colonna G)
            atd_val = parse_time_value(row_piano.get('ATD', None))
            if atd_val:
                scrivi_cella_sicura(ws_riepilogo, idx, 7, atd_val.strftime('%H:%M'))
            
            # Turno (colonna H)
            turno = str(row_piano.get('TURNO', '')).strip()
            if turno:
                scrivi_cella_sicura(ws_riepilogo, idx, 8, turno)
            
            # Fine Turno (colonna I) - se disponibile
            fine_turno = str(row_piano.get('FINE TURNO', '')).strip()
            if fine_turno:
                fine_turno_time = parse_time_value(fine_turno)
                if fine_turno_time:
                    scrivi_cella_sicura(ws_riepilogo, idx, 9, fine_turno_time.strftime('%H:%M'))
            
            # Cerca dati salvati per questo turno
            dati_turno = {}
            if dati_salvati and data_val:
                data_key = data_val.strftime("%Y-%m-%d")
                apt_val = str(row_piano.get('APT', '')).strip()
                
                if data_key and apt_val:
                    # Prova diverse chiavi
                    chiavi_possibili = []
                    if volo and volo.lower() not in ['nan', 'none', '']:
                        chiavi_possibili.append(f"{data_key}_{apt_val}_{volo.replace(' ', '_')}")
                    if std_val:
                        chiavi_possibili.append(f"{data_key}_{apt_val}_{std_val.strftime('%H%M')}")
                    chiavi_possibili.append(f"{data_key}_{apt_val}")
                    
                    for chiave in chiavi_possibili:
                        if chiave in dati_salvati:
                            dati_turno = dati_salvati[chiave]
                            break
            
            # Ore turno (colonna J) - dai dati salvati se disponibili
            if 'durata_effettiva_h' in dati_turno:
                durata_h = float(dati_turno.get('durata_effettiva_h', 0))
                if durata_h > 0:
                    ore_int = int(durata_h)
                    minuti_int = int((durata_h - ore_int) * 60)
                    if ore_int > 0:
                        scrivi_cella_sicura(ws_riepilogo, idx, 10, f"{ore_int}h {minuti_int}m")
                    else:
                        scrivi_cella_sicura(ws_riepilogo, idx, 10, f"{minuti_int}m")
            
            # Importo netto base (colonna K) - lascia vuoto, sarà compilato dall'assistente
            
            # Totale ore notte (colonna L) - formato testo se disponibile
            notte_min = int(dati_turno.get('notte_min', 0))
            if notte_min > 0:
                ore_notte = notte_min // 60
                min_notte = notte_min % 60
                if ore_notte > 0:
                    scrivi_cella_sicura(ws_riepilogo, idx, 12, f"{ore_notte}h {min_notte}m")
                else:
                    scrivi_cella_sicura(ws_riepilogo, idx, 12, f"{min_notte}min")
            
            # Ore extra (colonna N) - formato testo se disponibile
            extra_min = int(dati_turno.get('extra_min', 0))
            if extra_min > 0:
                ore_extra = extra_min // 60
                min_extra = extra_min % 60
                if ore_extra > 0:
                    scrivi_cella_sicura(ws_riepilogo, idx, 14, f"{ore_extra}h {min_extra}m")
                else:
                    scrivi_cella_sicura(ws_riepilogo, idx, 14, f"{min_extra}min")
            
            # % aggiuntiva / Ore notte in formato decimale (colonna M) - formula
            # Formula generica: l'assistente compilerà manualmente o da colonna L
            # Per ora lasciamo vuoto, l'assistente inserirà il valore decimale (es. 1.2+0.7)
            # Non generiamo formula automatica qui, l'assistente compilerà manualmente
            
            # Importo netto extra (colonna O) - formula generata dal file calcolo assistente
            # Determina la tariffa oraria extra corretta per l'aeroporto di questo turno
            apt_del_turno = str(row_piano.get('APT', '')).strip().upper()
            tipo_servizio_del_turno = str(row_piano.get('SERVIZIO', row_piano.get('SERVIZI', ''))).strip().upper()
            # Tariffa oraria extra corretta per aeroporto
            tariffa_extra_h = TARIFFE_EXTRA_PER_APT.get(apt_del_turno, 12.0)
            # Aggiustamenti per tipo servizio
            if apt_del_turno == 'FCO' and 'INCENTIVE' in tipo_servizio_del_turno:
                tariffa_extra_h = 15.0  # FCO incentive: €15/h
            elif apt_del_turno == 'NAP' and ('ARRIVI' in tipo_servizio_del_turno or 'MEET' in tipo_servizio_del_turno):
                tariffa_extra_h = 12.0  # NAP meet&greet: €12/h Senior
            elif apt_del_turno == 'NAP' and 'TRANSFER' in tipo_servizio_del_turno:
                tariffa_extra_h = 12.0  # NAP transfer: €12/h Senior
            
            if modulo_calcolo and hasattr(modulo_calcolo, 'genera_formula_excel_extra'):
                import inspect
                sig = inspect.signature(modulo_calcolo.genera_formula_excel_extra)
                if 'tariffa_extra_per_h' in sig.parameters:
                    # File calcolo aggiornato: supporta il parametro tariffa
                    formula_extra = modulo_calcolo.genera_formula_excel_extra(extra_min, tariffa_extra_h)
                else:
                    # File calcolo vecchio: chiama senza parametro (usa default hardcoded)
                    formula_extra = modulo_calcolo.genera_formula_excel_extra(extra_min)
            else:
                # Formula generica se non c'è file calcolo — usa tariffa corretta per APT
                tariff_str = int(tariffa_extra_h) if tariffa_extra_h == int(tariffa_extra_h) else tariffa_extra_h
                formula_extra = f"={tariff_str}/60*{extra_min if extra_min > 0 else 0}"
            
            try:
                cella_o = ws_riepilogo.cell(row=idx, column=15)
                cella_o.value = formula_extra
                # Rimuovi sfondo nero se presente
                if cella_o.fill and cella_o.fill.start_color:
                    rgb = str(cella_o.fill.start_color.rgb or '').upper()
                    if '000000' in rgb:
                        cella_o.fill = PatternFill()
            except:
                pass
            
            # Importo netto totale (colonna P) - formula generata dal file calcolo assistente
            # Verifica se ci sono ore notturne per includere colonna M
            ha_notte = notte_min > 0
            
            if modulo_calcolo and hasattr(modulo_calcolo, 'genera_formula_excel_totale'):
                # Passa anche il flag ha_notte se la funzione lo supporta
                try:
                    formula_totale = modulo_calcolo.genera_formula_excel_totale(idx, ha_notte=ha_notte)
                except TypeError:
                    # Se la funzione non accetta ha_notte, usa senza
                    formula_totale = modulo_calcolo.genera_formula_excel_totale(idx)
            else:
                # Formula generica basata sul template "DICEMBRE 2025"
                if ha_notte:
                    formula_totale = f"=SUM(O{idx}+M{idx}+K{idx})"
                else:
                    formula_totale = f"=SUM(O{idx}+K{idx})"
            
            try:
                cella_p = ws_riepilogo.cell(row=idx, column=16)
                cella_p.value = formula_totale
                # Rimuovi sfondo nero se presente
                if cella_p.fill and cella_p.fill.start_color:
                    rgb = str(cella_p.fill.start_color.rgb or '').upper()
                    if '000000' in rgb:
                        cella_p.fill = PatternFill()
            except:
                pass
            
            righe_dati.append(idx)
        
        # Genera formule di totale in base alle righe effettive
        if righe_dati:
            # Trova righe vuote dopo i dati per inserire i totali
            # Di solito c'è una riga vuota dopo i dati, poi i totali
            prima_riga_totale = max(righe_dati) + 2  # Prima riga dopo i dati + riga vuota
            
            # Formula totale parziale (se ci sono più gruppi)
            if len(righe_dati) > 0:
                riga_inizio_totale = min(righe_dati)
                riga_fine_totale = max(righe_dati)
                
                # Totale colonna P (importo netto totale)
                if modulo_calcolo and hasattr(modulo_calcolo, 'genera_formula_excel_totale_parziale'):
                    formula_totale_parziale = modulo_calcolo.genera_formula_excel_totale_parziale(
                        riga_inizio_totale, riga_fine_totale, 'P'
                    )
                else:
                    formula_totale_parziale = f"=SUM(P{riga_inizio_totale}:P{riga_fine_totale})"
                
                try:
                    cella_totale_p = ws_riepilogo.cell(row=prima_riga_totale, column=16)
                    cella_totale_p.value = formula_totale_parziale
                except:
                    pass
    
    # Genera nome file
    if output_path is None:
        nome_safe = nome_assistente.replace(" ", "_").replace("/", "_")
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = f"RIEPILOGO_ASSISTENZE_{nome_safe}_{timestamp}.xlsx"
    
    # Salva file
    wb_nuovo.save(output_path)
    
    return output_path

if __name__ == "__main__":
    # Test
    import sys
    if len(sys.argv) > 1:
        nome = sys.argv[1]
        output = genera_template_assistente(nome)
        print(f"Template generato: {output}")
    else:
        print("Uso: python genera_template_assistente.py <nome_assistente>")
