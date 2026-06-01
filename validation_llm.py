#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modulo di validazione LLM (Claude) per i file Excel del piano di lavoro.
Analizza struttura e contenuto e segnala problemi di formato/dati.
"""

import os
import json
import re
import pandas as pd
from typing import Optional
import streamlit as st


# ── Colonne attese nel formato 2026 ─────────────────────────────────────────
COLONNE_ATTESE = [
    "DATA", "CONVOCAZIONE", "AGENZIA", "TOUR OPERATOR",
    "SERVIZIO", "APT", "VOLO", "DEST.NE",
    "STA", "ATA", "STD", "ATD",
    "INIZIO TURNO", "FINE TURNO", "ASSISTENTE",
]

# ATD NON è obbligatoria (tutti i moduli hanno fallback su STD)
COLONNE_OBBLIGATORIE = [
    "DATA", "TOUR OPERATOR", "APT", "STD",
]


def _get_api_key() -> Optional[str]:
    """Recupera la API key da Streamlit secrets o variabile d'ambiente."""
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY")
        if key:
            return key
    except Exception:
        pass
    return os.environ.get("ANTHROPIC_API_KEY")


def _load_business_rules() -> str:
    """Carica le regole di business dal file business_rules.md."""
    base = os.path.dirname(os.path.abspath(__file__))
    rules_path = os.path.join(base, "business_rules.md")
    try:
        with open(rules_path, encoding="utf-8") as f:
            content = f.read()
        # Tronca a 8000 caratteri per non sforare il contesto
        return content[:8000]
    except Exception:
        return ""


# Regole per-TO: cosa serve per calcolare
# formato: {keyword_match: {"descr": ..., "critici": [(col, motivo)], "colonna_filtro": ..., "filtro_keyword": ...}}
TO_REQUISITI = {
    "veratour": {
        "descr": "Veratour (base 3h €75, extra ATD-fine×€18/h, notturno 23-05)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "usato come fallback se ATD manca, necessario per extra"),
            ("INIZIO TURNO", "orario inizio turno — serve per calcolo base"),
            ("FINE TURNO", "orario fine turno — serve per calcolo extra"),
        ],
        "utili": [("ATD", "orario decollo effettivo — calcola extra")],
    },
    "alpitour": {
        "descr": "Alpitour (base 3h, extra €20/h, notturno 23-06)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "necessario per calcolo base e extra"),
            ("INIZIO TURNO", "orario inizio turno"),
            ("FINE TURNO", "orario fine turno"),
        ],
        "utili": [("ATD", "orario decollo effettivo")],
    },
    "aliservice": {
        "descr": "Aliservice (agenzia, filtra su colonna AGENZIA, base 3h €55-€130)",
        "colonna_filtro": "AGENZIA",
        "critici": [
            ("INIZIO TURNO", "orario inizio (o CONVOCAZIONE per M&G)"),
            ("FINE TURNO", "orario fine turno"),
            ("STD", "usato come fine blocco se ATD manca"),
        ],
        "utili": [
            ("CONVOCAZIONE", "usata per M&G quando TURNO mancante"),
            ("ATD", "orario decollo effettivo per calcolo extra"),
        ],
    },
    "baobab": {
        "descr": "Baobab/TH (base 2h30, extra €18/h, notturno 22-06, festivo +30%)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "necessario per calcolo base e extra"),
            ("INIZIO TURNO", "orario inizio turno"),
            ("FINE TURNO", "orario fine turno"),
        ],
        "utili": [("ATD", "orario decollo effettivo per extra")],
    },
    "domina": {
        "descr": "Domina (base 2h30 per APT, extra €18/h, notturno 22-06, festivo +30%)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("APT", "determina la tariffa base: BGY €80, VRN €85, FCO €90, VCE €100"),
            ("STD", "usato come fine se ATD manca"),
            ("INIZIO TURNO", "orario inizio turno"),
            ("FINE TURNO", "orario fine turno"),
        ],
        "utili": [("ATD", "per extra se ritardo")],
    },
    "micheltours": {
        "descr": "MichelTours (base 3h, extra €18/h, notturno 22-06, festivo +30%)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "necessario"),
            ("INIZIO TURNO", "orario inizio turno"),
            ("FINE TURNO", "orario fine turno"),
        ],
        "utili": [("ATD", "per extra")],
    },
    "sand": {
        "descr": "SAND (fine turno=STD sempre, extra sempre €0, notturno da CONVOCAZIONE a STD)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "CRITICO: per Sand è la fine del turno (ATD ignorato contrattualmente)"),
            ("CONVOCAZIONE", "CRITICO: per Sand il notturno si calcola da CONVOCAZIONE a STD"),
            ("INIZIO TURNO", "orario inizio"),
        ],
        "utili": [],
    },
    "caboverdetime": {
        "descr": "Caboverdetime (base=CVC→STD, extra=STD→ATD, notturno 22-06)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("CONVOCAZIONE", "CRITICO: la base si calcola da CONVOCAZIONE a STD"),
            ("STD", "CRITICO: serve per base (CVC→STD) e come inizio extra"),
            ("ATD", "serve per il calcolo dell'extra (STD→ATD) — se manca extra=0"),
        ],
        "utili": [],
    },
    "rusconi": {
        "descr": "Rusconi (base fissa per APT: €100-€140, extra €20/h solo se ATD>STD)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("APT", "determina tariffa: BGY €110, FCO €115, VCE €140, altri €100"),
            ("STD", "necessario: extra calcolato solo se ATD>STD"),
            ("INIZIO TURNO", "orario inizio turno"),
            ("FINE TURNO", "orario fine turno"),
        ],
        "utili": [("ATD", "se ATD>STD viene calcolato l'extra €20/h")],
    },
    "iot": {
        "descr": "IOT (base 2h30, extra €18/h, notturno 22-06)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "necessario"),
            ("INIZIO TURNO", "orario inizio"),
            ("FINE TURNO", "orario fine"),
        ],
        "utili": [("ATD", "per extra")],
    },
    "flyness": {
        "descr": "Flyness (base 2h30, extra €20/h, notturno 22-06)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "necessario"),
            ("INIZIO TURNO", "orario inizio"),
            ("FINE TURNO", "orario fine"),
        ],
        "utili": [("ATD", "per extra")],
    },
    "rodocanachi": {
        "descr": "Rodocanachi (base=STD-2h30, extra €18/h da STD, notturno 22-06)",
        "colonna_filtro": "TOUR OPERATOR",
        "critici": [
            ("STD", "CRITICO: la base si calcola come STD-2h30"),
            ("APT", "determina tariffa: VCE/FCO €100, altri €90"),
        ],
        "utili": [("ATD", "per extra da STD")],
    },
}


def _analizza_dati_per_to(df: pd.DataFrame, col_map: dict) -> list:
    """
    Per ogni TO trovato nel file, verifica se i dati necessari
    per i suoi calcoli specifici sono presenti e valorizzati.
    Ritorna lista di problemi trovati.
    """
    problemi = []
    cols_upper = {str(c).strip().upper(): c for c in df.columns}

    def get_col(nome):
        return cols_upper.get(nome.upper())

    to_col = get_col("TOUR OPERATOR")
    agenzia_col = get_col("AGENZIA")
    conv_col = get_col("CONVOCAZIONE") or get_col("CONV.NE") or get_col("CONV")
    arrivi_col = get_col("ARRIVI/TRF") or get_col("ARRIVI TRF")

    for to_key, regole in TO_REQUISITI.items():
        # Trova righe di questo TO
        if regole["colonna_filtro"] == "AGENZIA" and agenzia_col:
            mask = df[agenzia_col].astype(str).str.contains(to_key, case=False, na=False)
        elif to_col:
            mask = df[to_col].astype(str).str.contains(to_key, case=False, na=False)
        else:
            continue

        df_to = df[mask]
        if df_to.empty:
            continue

        n_tot = len(df_to)

        # Controlla ogni campo critico
        for col_name, motivo in regole["critici"]:
            col_actual = get_col(col_name)
            if not col_actual:
                # La colonna non esiste nel file
                # Per INIZIO/FINE TURNO: accettabile se c'è TURNO
                turno_col = get_col("TURNO")
                if col_name in ("INIZIO TURNO", "FINE TURNO") and turno_col:
                    continue  # TURNO presente, ok
                problemi.append({
                    "to": to_key.upper(),
                    "tipo": "COLONNA_ASSENTE",
                    "colonna": col_name,
                    "n_righe": n_tot,
                    "n_vuoti": n_tot,
                    "motivo": motivo,
                    "msg": f"{to_key.upper()}: colonna '{col_name}' assente — {motivo}"
                })
            else:
                # La colonna esiste — quante righe sono vuote?
                vuoti = df_to[col_actual].isna().sum()
                # Per INIZIO/FINE TURNO: se c'è TURNO, le righe vuote sono normali (forward-fill)
                if col_name in ("INIZIO TURNO", "FINE TURNO"):
                    turno_col = get_col("TURNO")
                    if turno_col:
                        continue  # TURNO presente, forward-fill funziona
                if vuoti > 0:
                    pct = int(vuoti / n_tot * 100)
                    # Soglie diverse per criticità
                    # Campi CRITICI per calcolo: segnala anche pochi vuoti
                    problemi.append({
                        "to": to_key.upper(),
                        "tipo": "DATI_MANCANTI",
                        "colonna": col_name,
                        "n_righe": int(n_tot),
                        "n_vuoti": int(vuoti),
                        "pct": pct,
                        "motivo": motivo,
                        "msg": f"{to_key.upper()}: {vuoti}/{n_tot} righe ({pct}%) senza '{col_name}' — {motivo}"
                    })

        # Controllo speciale Aliservice M&G senza CONVOCAZIONE
        if to_key == "aliservice" and arrivi_col and conv_col:
            mg_mask = df_to[arrivi_col].astype(str).str.contains(r"M&G|MEET", case=False, na=False)
            df_mg = df_to[mg_mask]
            if not df_mg.empty:
                mg_senza_conv = df_mg[conv_col].isna().sum()
                if mg_senza_conv > 0:
                    problemi.append({
                        "to": "ALISERVICE",
                        "tipo": "MG_SENZA_CONVOCAZIONE",
                        "colonna": "CONVOCAZIONE",
                        "n_righe": len(df_mg),
                        "n_vuoti": int(mg_senza_conv),
                        "motivo": "Per M&G la CONVOCAZIONE è usata come orario inizio turno",
                        "msg": f"ALISERVICE: {mg_senza_conv} righe M&G senza CONVOCAZIONE — l'inizio del servizio non può essere determinato"
                    })

    return problemi


def _build_file_summary(file_path: str) -> dict:
    """
    Estrae un riassunto compatto del file Excel per il prompt LLM.
    Include analisi per-TO dei campi necessari ai calcoli specifici.
    """
    result = {
        "fogli": [],
        "foglio_analizzato": None,
        "colonne": [],
        "n_righe": 0,
        "colonne_mancanti": [],
        "campione": [],
        "anomalie": [],
        "valori_unici": {},
        "analisi_per_to": [],
    }

    try:
        xls = pd.ExcelFile(file_path)
        result["fogli"] = xls.sheet_names

        # Cerca foglio PIANO VOLI
        target = next(
            (s for s in xls.sheet_names if s.strip().upper() == "PIANO VOLI"),
            xls.sheet_names[0]
        )
        result["foglio_analizzato"] = target

        df = pd.read_excel(file_path, sheet_name=target)
        df.columns = [str(c).strip().upper() for c in df.columns]
        df = df.dropna(how="all")

        result["colonne"] = list(df.columns)
        result["n_righe"] = len(df)

        # Colonne obbligatorie mancanti (solo quelle che causano skip del foglio)
        colonne_skip = ["DATA", "APT"]
        result["colonne_mancanti"] = [c for c in colonne_skip if c not in df.columns]
        # Verifica has_orario: serve almeno (INIZIO TURNO + FINE TURNO) o TURNO
        has_orario = ("INIZIO TURNO" in df.columns and "FINE TURNO" in df.columns) or "TURNO" in df.columns
        if not has_orario:
            result["colonne_mancanti"].append("TURNO o (INIZIO TURNO + FINE TURNO)")

        # Campione: prime 8 righe non vuote
        key_cols = [c for c in COLONNE_ATTESE if c in df.columns][:10]
        sample_df = df[key_cols].head(8)
        result["campione"] = sample_df.astype(str).values.tolist()
        result["campione_header"] = key_cols

        # Valori unici per colonne categoriche
        for col in ["TOUR OPERATOR", "AGENZIA", "APT", "SERVIZIO"]:
            if col in df.columns:
                vals = df[col].dropna().astype(str).str.strip()
                vals = vals[~vals.str.lower().isin(["nan", "none", ""])]
                result["valori_unici"][col] = sorted(vals.unique().tolist())

        # Anomalie generali
        anomalie = []

        # Date in formato testo
        if "DATA" in df.columns:
            date_series = df["DATA"].dropna().astype(str)
            testo_dates = date_series[
                date_series.str.contains(r'[a-zA-Z\u00e0-\u00f9]', regex=True, na=False)
            ]
            if not testo_dates.empty:
                anomalie.append({
                    "tipo": "DATA_FORMATO_TESTO",
                    "n": len(testo_dates),
                    "esempi": testo_dates.head(3).tolist(),
                    "msg": f"{len(testo_dates)} righe con DATA in formato testo (es. '{testo_dates.iloc[0]}') — il programma le gestisce ma è meglio usare formato data Excel"
                })

        result["anomalie"] = anomalie

        # ── Analisi per-TO: controlla dati specifici per ogni calcolo ──────────
        result["analisi_per_to"] = _analizza_dati_per_to(df, {})

    except Exception as e:
        result["errore_lettura"] = str(e)

    return result


def _build_prompt(summary: dict, business_rules: str = "") -> str:
    """Costruisce il prompt da inviare a Claude, includendo le regole di business reali."""
    rules_section = ""
    if business_rules:
        rules_section = f"""
## REGOLE DI BUSINESS DEL PROGRAMMA (estratte dal codice sorgente reale)
QUESTE REGOLE SONO FONDAMENTALI: usale per evitare falsi allarmi.
{business_rules}
"""

    return f"""Sei un validatore esperto di file Excel per la gestione dei piani di lavoro aeroportuali di agenzie assistenti.

Analizza il seguente sommario di un file Excel e produci segnalazioni CONTESTUALI in italiano.
Devi conoscere bene le regole del programma per non segnalare cose che sono normali.
{rules_section}
## Sommario del file ricevuto
- Fogli presenti: {summary.get('fogli', [])}
- Foglio analizzato: {summary.get('foglio_analizzato')}
- Colonne presenti: {summary.get('colonne', [])}
- Numero righe dati: {summary.get('n_righe', 0)}
- Colonne obbligatorie mancanti: {summary.get('colonne_mancanti', [])}

## Tour Operator e Agenzie presenti nel file
{json.dumps(summary.get('valori_unici', {}), ensure_ascii=False, indent=2)}

## Anomalie rilevate automaticamente
{json.dumps(summary.get('anomalie', []), ensure_ascii=False, indent=2)}

## ANALISI PER TOUR OPERATOR — dati necessari per i calcoli
Questi sono i problemi concreti trovati analizzando riga per riga i dati necessari
per ogni specifico calcolo. OGNI problema qui è rilevante perché impatta il calcolo reale:
{json.dumps(summary.get('analisi_per_to', []), ensure_ascii=False, indent=2)}

## Campione dati (prime righe)
Intestazioni: {summary.get('campione_header', [])}
Righe:
{json.dumps(summary.get('campione', []), ensure_ascii=False, indent=2)}

## Istruzioni critiche
- I problemi in "ANALISI PER TOUR OPERATOR" sono GIÀ filtrati correttamente: segnalali tutti
- Ogni problema lì riportato significa che quel calcolo NON può essere eseguito o sarà impreciso
- ATD vuoto per TO che NON siano Caboverdetime: NON è un errore critico (fallback su STD)
- Caboverdetime senza ATD: AVVISO (extra non calcolabile)
- Righe singole con INIZIO/FINE TURNO vuoti: NON è un errore (forward-fill automatico)
- AGENZIA vuota per righe non-Aliservice: NON è un errore
- ASSISTENTE vuoto: NON è un errore
- TO gestiti da Aliservice (BRIXIA, FUTURA, ecc.): NON segnalare come problema
- TO senza modulo di calcolo: segnala solo come INFO

Rispondi SOLO con JSON valido:
{{
  "stato": "ok" | "attenzione" | "errore",
  "sommario": "frase breve sullo stato generale",
  "segnalazioni": [
    {{
      "gravita": "errore" | "avviso" | "info",
      "colonna": "nome colonna o null",
      "messaggio": "problema specifico in italiano",
      "suggerimento": "come correggerlo"
    }}
  ]
}}

Segnala SOLO problemi reali e rilevanti. Non includere nulla fuori dal JSON."""


def valida_file_con_llm(file_path: str) -> Optional[dict]:
    """
    Valida il file Excel con Claude.
    Inietta le regole di business dal codice sorgente reale nel prompt.
    Ritorna un dict con: stato, sommario, segnalazioni
    O None se l'API non è disponibile.
    """
    api_key = _get_api_key()
    if not api_key:
        return None

    try:
        import anthropic

        summary = _build_file_summary(file_path)
        business_rules = _load_business_rules()
        prompt = _build_prompt(summary, business_rules)

        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = message.content[0].text.strip()

        # Estrai JSON dalla risposta
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
        return None

    except Exception as e:
        return {"stato": "errore", "sommario": f"Errore validazione LLM: {str(e)}", "segnalazioni": []}


def render_validazione_llm(file_path: str, key_prefix: str = ""):
    """
    Componente Streamlit: mostra il pannello di validazione LLM.
    key_prefix: per evitare conflitti di chiave tra chiamate multiple.
    """
    api_key = _get_api_key()
    if not api_key:
        return  # Nessuna chiave, nessun pannello

    cache_key = f"_llm_validation_{file_path}_{key_prefix}"

    # Usa cache nel session_state per non richiamare ogni render
    if cache_key not in st.session_state:
        with st.spinner("🤖 Analisi intelligente del file in corso..."):
            result = valida_file_con_llm(file_path)
            st.session_state[cache_key] = result

    result = st.session_state.get(cache_key)
    if result is None:
        return

    stato = result.get("stato", "ok")
    sommario = result.get("sommario", "")
    segnalazioni = result.get("segnalazioni", [])

    # Header con colore per stato
    if stato == "errore":
        st.error(f"🤖 **Analisi AI** — {sommario}")
    elif stato == "attenzione":
        st.warning(f"🤖 **Analisi AI** — {sommario}")
    else:
        st.success(f"🤖 **Analisi AI** — {sommario}")

    if segnalazioni:
        for s in segnalazioni:
            gravita = s.get("gravita", "info")
            col_name = s.get("colonna")
            msg = s.get("messaggio", "")
            sug = s.get("suggerimento", "")

            col_label = f" `{col_name}`" if col_name else ""
            testo = f"**{col_label}** {msg}"
            if sug:
                testo += f"\n\n💡 *{sug}*"

            if gravita == "errore":
                st.error(testo)
            elif gravita == "avviso":
                st.warning(testo)
            else:
                st.info(testo)

    # Pulsante per rieseguire la validazione
    btn_key = f"riesegui_llm_{key_prefix}"
    if st.button("🔄 Riesegui analisi AI", key=btn_key, type="secondary"):
        if cache_key in st.session_state:
            del st.session_state[cache_key]
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# VERIFICA CALCOLI PER BLOCCO — Claude ricalcola indipendentemente
# ═══════════════════════════════════════════════════════════════════════════════

# Regole specifiche per TO da passare a Claude per il ricalcolo
TO_REGOLE_CALCOLO = {
    "veratour": """VERATOUR — Regole calcolo:
- Base: €75 per le prime 3 ore (turno fisso 3h)
- Extra: MAX(0, durata_effettiva - 180 min) × €18/h = €0.30/min
  - durata_effettiva = ATD - inizio_turno (se ATD presente, altrimenti STD - inizio_turno)
  - Se "NO DEC" nel turno: extra = 0
- Notturno: minuti nella fascia 23:00-05:00 × €0.083/min (≈ €5/h, differenziale 20% su €25/h)
- Festivo: (base + extra + notturno) × 1.20
- Totale = base + extra + notturno + festivo_maggiorazione""",

    "alpitour": """ALPITOUR — Regole calcolo:
- Base: per le prime 3 ore (tariffa base per APT, simile Veratour)
- Extra: MAX(0, durata_effettiva - 180 min) × €20/h
- Notturno: minuti 23:00-06:00 × €0.0625/min (15% = €3.75/h)
- Festivo: (base + extra) × 1.20
- Totale = base + extra + notturno + festivo_maggiorazione""",

    "aliservice": """ALISERVICE — Regole calcolo (AGENZIA che gestisce più TO):
- Tariffa base dipende da colonna SERVIZIO:
  - Tour Operator: €55 base + €15/h extra
  - MICE: €65 base + €15/h extra
  - Viaggi Studio: €55 + €15/h
  - VIP Service: €110 + €15/h
  - VIP Gate: €130 + €15/h
  - Meet & Greet (M&G in ARRIVI/TRF): €65 + €15/h
  - Default (nessun servizio): €55 + €15/h
- Durata base: 3 ore
- Extra: arrotondato per eccesso a 5 minuti
- Notturno: minuti 23:00-03:30 × €0.031/min
- Festivo: (base + extra + notturno) × 1.20""",

    "baobab": """BAOBAB (anche TH) — Regole calcolo:
- Base: tariffa per APT per 2h30 (150 min)
  - BGY: €80, VRN: €85, BLQ: €85, MXP: €90, altri: €90, VCE: €100
- Extra: MAX(0, durata - 150 min) × €18/h = €0.30/min
- Notturno: minuti 22:00-06:00 × tariffa_notte_per_min (≈ €0.107-0.133/min per APT)
- Festivo: (base + extra + notturno) × 1.30 (+30%)
- Totale = base + extra + notturno + festivo_maggiorazione""",

    "domina": """DOMINA — Regole calcolo:
- Base: tariffa per APT per 2h30 (150 min):
  - BGY: €80, VRN: €85, BLQ: €85, MXP: €90, NAP: €90, BRI: €90
  - CTA: €90, PMO: €90, FCO: €90, VCE: €100, PSI: €100
- Extra: MAX(0, durata_effettiva - 150 min) × €18/h
- Notturno: minuti 22:00-06:00 × tariffa_notte/min (20% della tariffa base oraria)
- Festivo: (base + extra + notturno) × 1.30 (+30%)
- Parcheggio: NAP €1/ora, BRI €6 ogni 3h (eccesso)
- Totale = base + extra + notturno + festivo_magg. + parcheggio""",

    "micheltours": """MICHELTOURS — Regole calcolo:
- Base: tariffa per APT per 3h (180 min):
  - BGY: €85, MXP: €90, VCE: €90
- Extra: MAX(0, durata - 180 min) × €18/h
- Notturno: minuti 22:00-06:00 × tariffa_notte/min
- Festivo: totale × 1.30
- Totale = base + extra + notturno + festivo_magg.""",

    "sand": """SAND — Regole calcolo SPECIALI (contratto diverso):
- Base: da INIZIO TURNO a STD (NON ATD — ATD viene ignorato contrattualmente)
- Extra: SEMPRE €0 (nessun extra per contratto)
- Notturno: minuti 22:00-03:59 nell'intervallo da CONVOCAZIONE a STD
- Festivo: (base + notturno) × 1.30
- Tariffa base per APT: BGY: €65, VRN/BLQ/NAP: €70, VCE: €75, FCO: €77
- ATTENZIONE: la fine del turno è sempre STD, non ATD""",

    "caboverdetime": """CABOVERDETIME — Regole calcolo SPECIALI:
- Base: da CONVOCAZIONE (CVC) a STD (durata variabile, NON fissa)
  - Importo base proporzionale alla durata CVC→STD
- Extra: da STD a ATD (se ATD presente, altrimenti 0)
  - Extra = (ATD - STD) in minuti × €18/h
- Notturno: minuti 22:00-06:00 sull'intera durata CVC→ATD (o CVC→STD se no ATD)
- Festivo: totale × 1.30
- ATTENZIONE: CONVOCAZIONE è fondamentale per il calcolo della base""",

    "rusconi": """RUSCONI — Regole calcolo:
- Base: tariffa fissa per APT per 2h30 (150 min):
  - BGY: €110, FCO: €115, VCE: €140, tutti gli altri: €100
- Extra: MAX(0, ATD - STD) × €20/h — solo se ATD > STD
  - Se ATD manca o ATD ≤ STD: extra = 0
- Notturno: minuti 22:00-06:00 × tariffa/min
- Festivo: totale × 1.30
- Totale = base + extra + notturno + festivo_magg.""",

    "iot": """IOT — Regole calcolo:
- Base: tariffa per APT per 2h30, extra €18/h, notturno 22:00-06:00, festivo +30%""",

    "flyness": """FLYNESS — Regole calcolo:
- Base: tariffa per APT per 2h30, extra €20/h, notturno 22:00-06:00, festivo +30%""",

    "rodocanachi": """RODOCANACHI (IOT Viaggi) — Regole calcolo:
- Base: convocazione = STD - 2h30, tariffa per APT (VCE/FCO: €100, altri: €90)
- Extra: da STD × €18/h
- Notturno: 22:00-06:00, festivo +30%""",
}


def _get_to_regole(to_name: str) -> str:
    """Restituisce le regole di calcolo per il TO specificato."""
    to_lower = to_name.lower().strip()
    for key, regole in TO_REGOLE_CALCOLO.items():
        if key in to_lower:
            return regole
    return f"TO '{to_name}': regole non codificate, usa logica standard (base 2h30, extra €18/h)."


def _safe_num(val) -> Optional[float]:
    """Converte valore in float, ritorna None se non possibile."""
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        return float(val)
    except Exception:
        return None


def _build_batch_prompt(to_name: str, rows: list) -> str:
    """
    Costruisce il prompt per Claude per ricalcolare una batch di blocchi.
    rows: lista di dict con i campi del blocco.
    """
    regole = _get_to_regole(to_name)
    rows_json = json.dumps(rows, ensure_ascii=False, default=str, indent=2)

    return f"""Sei un calcolatore esperto di tariffe per servizi aeroportuali SCAY Group.

## TOUR OPERATOR: {to_name.upper()}

## Regole di calcolo da applicare:
{regole}

## Blocchi da calcolare:
{rows_json}

## Istruzioni:
Per ogni blocco, calcola indipendentemente i valori e fornisci una spiegazione del calcolo.
Usa i campi: DATA, APT, INIZIO_TURNO, FINE_TURNO, STD, ATD, CONVOCAZIONE, DURATA_MIN, FESTIVO.

Rispondi SOLO con JSON array valido, un oggetto per ogni blocco nella stessa sequenza:
[
  {{
    "idx": 0,
    "base_ai": 75.00,
    "extra_min_ai": 45,
    "extra_ai": 13.50,
    "notturno_min_ai": 0,
    "notturno_ai": 0.00,
    "festivo_ai": 0.00,
    "totale_ai": 88.50,
    "calcolo_ai": "Base 3h €75.00 | Extra 45min×€18/h=€13.50 | Nott. 0min | Totale €88.50"
  }}
]

Il campo "calcolo_ai" deve essere una riga di testo concisa che mostra i passaggi.
Non includere nulla al di fuori del JSON."""


def verifica_calcoli_con_llm(detail_df: pd.DataFrame, max_rows_per_to: int = 30) -> Optional[pd.DataFrame]:
    """
    Per ogni TO nel detail_df, chiede a Claude di ricalcolare ogni blocco
    indipendentemente usando le regole del TO.
    Ritorna il DataFrame con colonne aggiuntive AI.
    max_rows_per_to: quante righe per TO mandare a Claude (evita token overflow).
    """
    api_key = _get_api_key()
    if not api_key or detail_df is None or detail_df.empty:
        return None

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
    except Exception:
        return None

    # Colonne AI da aggiungere
    result_df = detail_df.copy()
    result_df["BASE AI (€)"] = None
    result_df["EXTRA AI (€)"] = None
    result_df["NOTTURNO AI (€)"] = None
    result_df["TOTALE AI (€)"] = None
    result_df["CALCOLO AI"] = ""
    result_df["DIFF SOFTWARE vs AI"] = ""

    # Normalizza nomi colonne per lookup flessibile
    cols = {str(c).strip().upper(): c for c in detail_df.columns}

    def gc(nome):
        """get column actual name, case-insensitive."""
        return cols.get(nome.upper())

    # Identifica colonna TOUR OPERATOR
    to_col = gc("TOUR OPERATOR") or gc("TO")
    if not to_col:
        return None

    # Gruppa per TO
    to_groups = detail_df[to_col].astype(str).str.strip().str.upper().unique()

    for to_name in to_groups:
        if to_name.lower() in ("nan", "none", ""):
            continue

        to_mask = detail_df[to_col].astype(str).str.strip().str.upper() == to_name
        to_rows_idx = detail_df[to_mask].index.tolist()

        if not to_rows_idx:
            continue

        # Limita a max_rows_per_to per evitare token overflow
        chunk_idx = to_rows_idx[:max_rows_per_to]
        chunk_df = detail_df.loc[chunk_idx]

        # Prepara batch di righe per Claude
        def safe_str(v):
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass
            return str(v) if v is not None else None

        rows_for_claude = []
        for i, (idx, row) in enumerate(chunk_df.iterrows()):
            r = {"idx": i, "row_index": idx}
            for campo, keys in {
                "DATA": ["DATA"],
                "APT": ["APT"],
                "INIZIO_TURNO": ["INIZIO TURNO", "TURNO"],
                "FINE_TURNO": ["FINE TURNO"],
                "STD": ["STD", "STD_SCELTO"],
                "ATD": ["ATD", "ATD_SCELTO"],
                "CONVOCAZIONE": ["CONVOCAZIONE", "CONV.NE", "CONV"],
                "DURATA_MIN": ["DURATA", "DURATA_TURNO_MIN"],
                "FESTIVO": ["FESTIVO"],
                "ASSISTENTE": ["ASSISTENTE"],
            }.items():
                for k in keys:
                    col_actual = gc(k)
                    if col_actual and col_actual in row.index:
                        r[campo] = safe_str(row[col_actual])
                        break
                else:
                    r[campo] = None
            rows_for_claude.append(r)

        if not rows_for_claude:
            continue

        try:
            prompt = _build_batch_prompt(to_name, rows_for_claude)
            message = client.messages.create(
                model="claude-opus-4-5",
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = message.content[0].text.strip()

            # Estrai JSON array
            json_match = re.search(r'\[.*\]', raw, re.DOTALL)
            if not json_match:
                continue

            ai_results = json.loads(json_match.group())

            # Mappa risultati AI al DataFrame
            for ai_row in ai_results:
                i = ai_row.get("idx")
                if i is None or i >= len(rows_for_claude):
                    continue
                actual_idx = rows_for_claude[i]["row_index"]

                base_ai = _safe_num(ai_row.get("base_ai"))
                extra_ai = _safe_num(ai_row.get("extra_ai"))
                nott_ai = _safe_num(ai_row.get("notturno_ai"))
                tot_ai = _safe_num(ai_row.get("totale_ai"))
                calcolo = ai_row.get("calcolo_ai", "")

                result_df.at[actual_idx, "BASE AI (€)"] = base_ai
                result_df.at[actual_idx, "EXTRA AI (€)"] = extra_ai
                result_df.at[actual_idx, "NOTTURNO AI (€)"] = nott_ai
                result_df.at[actual_idx, "TOTALE AI (€)"] = tot_ai
                result_df.at[actual_idx, "CALCOLO AI"] = calcolo

                # Confronta con software
                tot_sw_col = gc("TOTALE_EUR") or gc("TOTALE (€)") or gc("TOTALE")
                if tot_sw_col and tot_ai is not None:
                    try:
                        tot_sw = _safe_num(detail_df.at[actual_idx, tot_sw_col])
                        if tot_sw is not None:
                            diff = abs(tot_sw - tot_ai)
                            if diff < 0.05:
                                result_df.at[actual_idx, "DIFF SOFTWARE vs AI"] = "✅ Uguale"
                            else:
                                result_df.at[actual_idx, "DIFF SOFTWARE vs AI"] = f"⚠️ Diff €{diff:.2f} (SW={tot_sw:.2f}, AI={tot_ai:.2f})"
                    except Exception:
                        pass

        except Exception:
            continue

    return result_df


def render_verifica_calcoli_llm(detail_df: pd.DataFrame):
    """
    Componente Streamlit: bottone per avviare la verifica calcoli con Claude.
    Mostra il DataFrame arricchito e il download Excel con colonne AI.
    """
    api_key = _get_api_key()
    if not api_key or detail_df is None or detail_df.empty:
        return

    st.markdown("---")
    st.markdown("### 🤖 Verifica Calcoli con AI")
    st.caption(
        "Claude ricalcola ogni blocco indipendentemente usando le regole reali di ogni Tour Operator "
        "e confronta il risultato con quello del software."
    )

    cache_key = "_llm_verifica_calcoli"

    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        if st.button("🤖 Avvia verifica calcoli AI", type="primary", key="avvia_verifica_ai"):
            if cache_key in st.session_state:
                del st.session_state[cache_key]
            with st.spinner("Claude sta ricalcolando ogni blocco... (può richiedere 30-60 secondi)"):
                verified_df = verifica_calcoli_con_llm(detail_df)
                st.session_state[cache_key] = verified_df

    with col_info:
        st.info(
            "💡 Il calcolo AI è indipendente dal software. "
            "Se i totali differiscono, clicca sul blocco per vedere il dettaglio."
        )

    verified_df = st.session_state.get(cache_key)
    if verified_df is None:
        return

    # Mostra statistiche confronto
    diff_col = "DIFF SOFTWARE vs AI"
    if diff_col in verified_df.columns:
        n_ok = (verified_df[diff_col] == "✅ Uguale").sum()
        n_diff = verified_df[diff_col].str.startswith("⚠️").sum()
        n_tot = len(verified_df[verified_df[diff_col] != ""])

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Blocchi verificati", n_tot)
        with c2:
            st.metric("✅ Concordanti", n_ok)
        with c3:
            st.metric("⚠️ Differenze", n_diff)

    # Mostra solo colonne rilevanti + AI
    cols_to_show = []
    cols_upper = {str(c).strip().upper(): c for c in verified_df.columns}
    for nome in ["DATA", "APT", "TOUR OPERATOR", "INIZIO TURNO", "FINE TURNO", "STD", "ATD",
                 "TURNO_EUR", "TURNO (€)", "BASE_EUR", "EXTRA_EUR", "EXTRA (€)",
                 "NOTTURNO_EUR", "NOTTURNO (€)", "TOTALE_EUR", "TOTALE (€)",
                 "BASE AI (€)", "EXTRA AI (€)", "NOTTURNO AI (€)", "TOTALE AI (€)",
                 "CALCOLO AI", "DIFF SOFTWARE vs AI"]:
        col_actual = cols_upper.get(nome.upper())
        if col_actual and col_actual not in cols_to_show:
            cols_to_show.append(col_actual)

    if cols_to_show:
        st.dataframe(verified_df[cols_to_show], use_container_width=True, hide_index=True)

    # Download Excel con colonne AI
    try:
        import io
        from openpyxl.styles import PatternFill, Font
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            verified_df.to_excel(writer, sheet_name="DettaglioConAI", index=False)
            # Evidenzia righe con differenze
            ws = writer.sheets["DettaglioConAI"]
            diff_col_idx = None
            for col_idx, col_name in enumerate(verified_df.columns, 1):
                if col_name == "DIFF SOFTWARE vs AI":
                    diff_col_idx = col_idx
                    break
            if diff_col_idx:
                fill_warn = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                fill_ok = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                for row_idx, val in enumerate(verified_df["DIFF SOFTWARE vs AI"], 2):
                    cell = ws.cell(row=row_idx, column=diff_col_idx)
                    if str(val).startswith("⚠️"):
                        for c in range(1, len(verified_df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = fill_warn
                    elif str(val).startswith("✅"):
                        ws.cell(row=row_idx, column=diff_col_idx).fill = fill_ok
        buf.seek(0)
        st.download_button(
            "⬇️ Scarica Excel con verifica AI",
            data=buf.getvalue(),
            file_name="dettaglio_con_verifica_ai.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_ai_verify",
        )
    except Exception as e:
        st.warning(f"Download non disponibile: {e}")

